#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
呼び出しツリー可視化スクリプト
TSVファイルから呼び出しツリーを生成します
"""

import csv
import json
import re
import sys
from collections import defaultdict
from typing import Dict, List, Optional, Set

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

# Git Bash上でパイプを使うと、stdoutがCP932として扱われるのを防ぐ
sys.stdout.reconfigure(encoding="utf-8")


class ExclusionRuleManager:
    """除外ルールを管理するクラス"""

    def __init__(self, exclusion_file: Optional[str] = None):
        """
        コンストラクタ

        Args:
            exclusion_file: 除外ルールファイルのパス（Noneの場合はデフォルトファイルを使用）
        """
        self.include_exclusions: Set[str] = set()  # Iモード: 対象自体を除外
        self.exclude_children: Set[str] = set()  # Eモード: 配下のみ除外

        # デフォルトファイル名
        if exclusion_file is None:
            exclusion_file = "exclusion_rules.txt"

        # ファイルが存在する場合のみ読み込む
        import os

        if os.path.exists(exclusion_file):
            self.load_rules(exclusion_file)

    def load_rules(self, file_path: str) -> None:
        """
        除外ルールをファイルから読み込む

        Args:
            file_path: 除外ルールファイルのパス
        """
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith("#"):
                        # 空行とコメント行はスキップ
                        continue

                    parts = line.split("\t")
                    if len(parts) != 2:
                        print(
                            f"警告: 行 {line_num} のフォーマットが不正です: {line}",
                            file=sys.stderr,
                        )
                        continue

                    target, mode = parts
                    target = target.strip()
                    mode = mode.strip().upper()

                    if mode == "I":
                        self.include_exclusions.add(target)
                    elif mode == "E":
                        self.exclude_children.add(target)
                    else:
                        print(
                            f"警告: 行 {line_num} の除外モードが不正です: {mode} (I または E を指定してください)",
                            file=sys.stderr,
                        )

            print(f"除外ルールを読み込みました: {file_path}", file=sys.stderr)
            print(
                f"  Iモード(対象除外): {len(self.include_exclusions)} 件",
                file=sys.stderr,
            )
            print(
                f"  Eモード(配下除外): {len(self.exclude_children)} 件", file=sys.stderr
            )

        except Exception as e:
            print(f"除外ルールファイルの読み込みに失敗しました: {e}", file=sys.stderr)

    def should_include(self, method_or_class: str) -> bool:
        """
        メソッド/クラスがIモード（対象自体を除外）の対象かチェック

        Args:
            method_or_class: メソッドシグネチャまたはクラス名

        Returns:
            True: 表示すべき（除外対象ではない）
            False: 除外すべき（除外対象）
        """
        # メソッドシグネチャ全体が除外対象か
        if method_or_class in self.include_exclusions:
            return False

        # クラス名を抽出して除外対象かチェック
        class_name = self._extract_class_name(method_or_class)
        if class_name and class_name in self.include_exclusions:
            return False

        # クラス名を除くメソッド部分のみを抽出して除外対象かチェック
        method_part = self._extract_method_part(method_or_class)
        if method_part and method_part in self.include_exclusions:
            return False

        return True

    def should_exclude_children(self, method_or_class: str) -> bool:
        """
        メソッド/クラスがEモード（配下のみ除外）の対象かチェック

        Args:
            method_or_class: メソッドシグネチャまたはクラス名

        Returns:
            True: 配下を除外すべき
            False: 配下も展開すべき
        """
        # メソッドシグネチャ全体が除外対象か
        if method_or_class in self.exclude_children:
            return True

        # クラス名を抽出して除外対象かチェック
        class_name = self._extract_class_name(method_or_class)
        if class_name and class_name in self.exclude_children:
            return True

        # クラス名を除くメソッド部分のみを抽出して除外対象かチェック
        method_part = self._extract_method_part(method_or_class)
        if method_part and method_part in self.exclude_children:
            return True

        return False

    def _extract_class_name(self, method_signature: str) -> Optional[str]:
        """
        メソッドシグネチャからクラス名を抽出

        Args:
            method_signature: メソッドシグネチャ (例: "com.example.Class#method()")

        Returns:
            クラス名 (例: "com.example.Class")、抽出できない場合はNone
        """
        if "#" in method_signature:
            return method_signature.split("#")[0]
        return None

    def _extract_method_part(self, method_signature: str) -> Optional[str]:
        """
        メソッドシグネチャからクラスを除去したメソッド部分のみを抽出

        Args:
            method_signature: メソッドシグネチャ (例: "com.example.Class#method()")

        Returns:
            クラスを除去したメソッド部分 (例: "method()")、抽出できない場合はNone
        """
        if "#" in method_signature:
            method_part = method_signature.split("#")[1]
            # # 引数部分を除去
            # if "(" in method_part:
            #     return method_part.split("(")[0]
            return method_part
        return None


class CallTreeVisualizer:
    def __init__(
        self,
        input_file: str,
        exclusion_file: Optional[str] = None,
        output_tsv_encoding: str = "Shift_JIS",
        debug_mode: bool = False,  # デバッグモード
    ):
        """
        コンストラクタ

        Args:
            input_file: 入力ファイルのパス（JSONまたはTSV）
            exclusion_file: 除外ルールファイルのパス（Noneの場合はデフォルト）
            output_tsv_encoding: 出力TSVファイルのエンコーディング
            debug_mode: デバッグモード（Trueの場合、インスタンス収集情報を出力）
        """
        self.input_file: str = input_file
        self.forward_calls: Dict[str, List[Dict[str, str]]] = defaultdict(list)
        self.reverse_calls: Dict[str, List[str]] = defaultdict(list)
        self.method_info: Dict[str, Dict[str, Optional[str]]] = {}
        self.class_info: Dict[str, List[str]] = {}
        # クラス・インターフェースのメタデータ（annotations, javadoc等）
        self.class_data: Dict[str, Dict] = (
            {}
        )  # className -> {annotations, javadoc, superClass, directInterfaces, allInterfaces}
        self.interface_data: Dict[str, Dict] = (
            {}
        )  # interfaceName -> {annotations, javadoc, superInterfaces}
        self.exclusion_manager: ExclusionRuleManager = ExclusionRuleManager(
            exclusion_file
        )
        self.output_tsv_encoding: str = output_tsv_encoding
        self.debug_mode: bool = debug_mode
        self.load_data()

    def load_data(self):
        """入力ファイルを読み込む（JSON/TSV自動判定）"""
        if self.input_file.lower().endswith(".json"):
            self._load_json_data()
        else:
            self._load_tsv_data()

    def _load_json_data(self):
        """JSON形式（統合形式）からデータを読み込む"""
        with open(self.input_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        methods = data.get("methods", [])
        for method in methods:
            method_sig = method.get("method", "")
            if not method_sig:
                continue

            class_name = method.get("class", "")
            parent_classes_str = method.get("parentClasses", "")

            # メソッド情報を保存
            self.method_info[method_sig] = {
                "class": class_name,
                "parent": parent_classes_str,  # 親クラス情報（カンマ区切り文字列）
                "visibility": method.get("visibility", ""),
                "is_static": method.get("isStatic", False),
                "is_entry_point": method.get("isEntryPoint", False),
                "entry_type": method.get("entryType", ""),
                "annotations": ",".join(method.get("annotations", [])),
                "class_annotations": ",".join(method.get("classAnnotations", [])),
                "sql": " ||| ".join(method.get("sqlStatements", [])),
                "javadoc": method.get("javadoc", ""),
                "class_javadoc": method.get("classJavadoc", ""),
                "hit_words": ",".join(method.get("hitWords", [])),
                "createdInstances": method.get(
                    "createdInstances", []
                ),  # 生成されたインスタンス
                "httpCalls": method.get("httpCalls", []),  # HTTPクライアント呼び出し
            }

            # クラス階層情報を保存（parentClassesから取得した全親クラス・インターフェース）
            if class_name and parent_classes_str:
                parents = [
                    p.strip() for p in parent_classes_str.split(",") if p.strip()
                ]
                # 既存の情報がない場合、または新しい情報がある場合は更新
                if class_name not in self.class_info or not self.class_info[class_name]:
                    self.class_info[class_name] = parents
                else:
                    # 既存のリストに新しい親を追加
                    existing = set(self.class_info[class_name])
                    for p in parents:
                        if p not in existing:
                            self.class_info[class_name].append(p)

            # 呼び出し関係を保存（callsから詳細情報を取得）
            calls_data = method.get("calls", [])
            for call_item in calls_data:
                # 新形式：オブジェクト配列
                if isinstance(call_item, dict):
                    callee = call_item.get("method", "")
                    if callee:
                        is_parent = (
                            "Yes" if call_item.get("isParentMethod", False) else "No"
                        )
                        impls = call_item.get("implementations", "")
                        self.forward_calls[method_sig].append(
                            {
                                "method": callee,
                                "is_parent_method": is_parent,
                                "implementations": impls,
                            }
                        )
                else:
                    # 後方互換性：文字列配列
                    self.forward_calls[method_sig].append(
                        {
                            "method": call_item,
                            "is_parent_method": "No",
                            "implementations": "",
                        }
                    )

            # 逆引き呼び出し関係を保存
            for caller in method.get("calledBy", []):
                self.reverse_calls[method_sig].append(caller)

        # classesセクションを読み込み
        classes = data.get("classes", [])
        for cls in classes:
            class_name = cls.get("className", "")
            if class_name:
                self.class_data[class_name] = {
                    "annotations": cls.get("annotations", []),
                    "annotationRaws": cls.get("annotationRaws", []),
                    "javadoc": cls.get("javadoc", ""),
                    "superClass": cls.get("superClass", ""),
                    "directInterfaces": cls.get("directInterfaces", []),
                    "allInterfaces": cls.get("allInterfaces", []),
                    "fieldInitializers": cls.get(
                        "fieldInitializers", []
                    ),  # フィールド初期化情報
                }

        # interfacesセクションを読み込み
        interfaces = data.get("interfaces", [])
        for iface in interfaces:
            iface_name = iface.get("interfaceName", "")
            if iface_name:
                self.interface_data[iface_name] = {
                    "annotations": iface.get("annotations", []),
                    "annotationRaws": iface.get("annotationRaws", []),
                    "javadoc": iface.get("javadoc", ""),
                    "superInterfaces": iface.get("superInterfaces", []),
                }

    def _get_all_class_annotations(self, class_name: str) -> List[str]:
        """クラスとその親クラス・インターフェースのすべてのアノテーションを再帰的に取得"""
        if not class_name:
            return []

        all_annotations: List[str] = []
        visited: Set[str] = set()
        self._collect_annotations_recursive(class_name, all_annotations, visited)
        return all_annotations

    def _collect_annotations_recursive(
        self, type_name: str, all_annotations: List[str], visited: Set[str]
    ) -> None:
        """アノテーションを再帰的に収集"""
        if not type_name or type_name in visited:
            return
        visited.add(type_name)

        # クラスとして検索
        if type_name in self.class_data:
            cls = self.class_data[type_name]
            for ann in cls.get("annotations", []):
                if ann not in all_annotations:
                    all_annotations.append(ann)
            # 親クラスを辿る
            super_class = cls.get("superClass", "")
            if super_class:
                self._collect_annotations_recursive(
                    super_class, all_annotations, visited
                )
            # インターフェースを辿る
            for iface in cls.get("directInterfaces", []):
                self._collect_annotations_recursive(iface, all_annotations, visited)

        # インターフェースとして検索
        if type_name in self.interface_data:
            iface = self.interface_data[type_name]
            for ann in iface.get("annotations", []):
                if ann not in all_annotations:
                    all_annotations.append(ann)
            # 親インターフェースを辿る
            for super_iface in iface.get("superInterfaces", []):
                self._collect_annotations_recursive(
                    super_iface, all_annotations, visited
                )

    def _get_class_javadoc(self, class_name: str) -> str:
        """クラスのJavadocを取得（なければ空文字）"""
        if class_name in self.class_data:
            return self.class_data[class_name].get("javadoc", "")
        if class_name in self.interface_data:
            return self.interface_data[class_name].get("javadoc", "")
        return ""

    def _get_all_class_annotation_raws(self, class_name: str) -> List[str]:
        """クラスとその親クラス・インターフェースのフル形式アノテーションを再帰的に取得

        エンドポイントパス抽出用にannotationRaws（@RequestMapping(path = "/bill")形式）を返す
        """
        if not class_name:
            return []

        all_annotations: List[str] = []
        visited: Set[str] = set()
        self._collect_annotation_raws_recursive(class_name, all_annotations, visited)
        return all_annotations

    def _collect_annotation_raws_recursive(
        self, type_name: str, all_annotations: List[str], visited: Set[str]
    ) -> None:
        """フル形式アノテーションを再帰的に収集"""
        if not type_name or type_name in visited:
            return
        visited.add(type_name)

        # クラスとして検索
        if type_name in self.class_data:
            cls = self.class_data[type_name]
            for ann in cls.get("annotationRaws", []):
                if ann not in all_annotations:
                    all_annotations.append(ann)
            # 親クラスを辿る
            super_class = cls.get("superClass", "")
            if super_class:
                self._collect_annotation_raws_recursive(
                    super_class, all_annotations, visited
                )
            # インターフェースを辿る
            for iface in cls.get("directInterfaces", []):
                self._collect_annotation_raws_recursive(iface, all_annotations, visited)

        # インターフェースとして検索
        if type_name in self.interface_data:
            iface = self.interface_data[type_name]
            for ann in iface.get("annotationRaws", []):
                if ann not in all_annotations:
                    all_annotations.append(ann)
            # 親インターフェースを辿る
            for super_iface in iface.get("superInterfaces", []):
                self._collect_annotation_raws_recursive(
                    super_iface, all_annotations, visited
                )

    def _load_tsv_data(self):
        """TSV形式からデータを読み込む（後方互換性）"""
        with open(self.input_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for row in reader:
                caller = row["呼び出し元メソッド"]
                callee = row["呼び出し先メソッド"]
                direction = row["方向"]

                # メソッド情報を保存
                if caller and (
                    caller not in self.method_info
                    or self.method_info[caller]["visibility"] == ""
                ):
                    if caller not in self.method_info:
                        self.method_info[caller] = {}
                    self.method_info[caller] |= {
                        "class": row["呼び出し元クラス"],
                        "parent": row["呼び出し元の親クラス"],
                        "visibility": row.get("可視性", ""),
                        "is_static": row.get("Static", "") == "Yes",
                        "is_entry_point": row.get("エントリーポイント候補", "")
                        == "Yes",
                        "entry_type": row.get("エントリータイプ", ""),
                        "annotations": row.get("アノテーション", ""),
                        "class_annotations": row.get("クラスアノテーション", ""),
                    }

                    # クラス階層情報を保存
                    if row["呼び出し元クラス"]:
                        parents = [
                            p.strip()
                            for p in row["呼び出し元の親クラス"].split(",")
                            if p.strip()
                        ]
                        self.class_info[row["呼び出し元クラス"]] = parents

                if callee:
                    if callee not in self.method_info:
                        self.method_info[callee] = {
                            "class": row["呼び出し先クラス"],
                            "parent": "",
                            "sql": (
                                row.get("SQL文", "") if direction == "Forward" else ""
                            ),
                            "visibility": "",
                            "is_static": False,
                            "is_entry_point": False,
                            "annotations": "",
                            "javadoc": (
                                row.get("メソッドJavadoc", "")
                                if direction == "Forward"
                                else ""
                            ),
                        }
                    else:
                        if not self.method_info[callee].get("sql"):
                            # SQL文は呼び出し先の情報に基づく
                            self.method_info[callee] |= {"sql": row.get("SQL文", "")}
                        if not self.method_info[callee].get("javadoc"):
                            self.method_info[callee] |= {
                                "javadoc": row.get("メソッドJavadoc", "")
                            }

                    # クラス階層情報を保存（呼び出し先）
                    callee_class = row["呼び出し先クラス"]
                    callee_parents_str = row.get("呼び出し先の親クラス", "")
                    if callee_class:
                        parents = [
                            p.strip()
                            for p in callee_parents_str.split(",")
                            if p.strip()
                        ]
                        if (
                            callee_class not in self.class_info
                            or not self.class_info[callee_class]
                        ):
                            self.class_info[callee_class] = parents

                # 呼び出し関係を保存
                if direction == "Forward" and caller and callee:
                    self.forward_calls[caller].append(
                        {
                            "method": callee,
                            "is_parent_method": row["呼び出し先は親クラスのメソッド"],
                            "implementations": row["呼び出し先の実装クラス候補"],
                        }
                    )
                elif direction == "Reverse" and caller and callee:
                    self.reverse_calls[caller].append(callee)

    def print_forward_tree(
        self,
        root_method: str,
        max_depth: int = 50,
        show_class: bool = False,
        show_sql: bool = False,
        follow_implementations: bool = True,
        verbose: bool = False,
        use_tab: bool = False,
        short_mode: bool = False,
    ):
        """呼び出し元からのツリーを表示

        Args:
            root_method: 起点メソッド
            max_depth: 最大深度
            show_class: クラス情報を表示するか
            show_sql: SQL情報を表示するか
            follow_implementations: 実装クラス候補がある場合、それも追跡するか
            verbose: 詳細表示（Javadocを表示）
            use_tab: Trueの場合、ハードタブでインデントし、プレフィックスを省略
            short_mode: Trueの場合、クラス名からパッケージ名を省いて表示
        """
        print(f"\n{'=' * 80}")
        print(f"呼び出しツリー (起点: {root_method})")
        print(f"{'=' * 80}\n")

        visited: set[str] = set()
        # 最大深度到達フラグを初期化
        max_depth_reached: List[bool] = [False]

        self._print_tree_recursive(
            root_method,
            0,
            max_depth,
            visited,
            show_class,
            show_sql,
            is_forward=True,
            follow_implementations=follow_implementations,
            verbose=verbose,
            use_tab=use_tab,
            short_mode=short_mode,
            accumulated_instances=None,  # ルートから累積開始
            max_depth_reached=max_depth_reached,
        )

        # 最大深度に到達した場合の警告を出力
        if max_depth_reached[0]:
            print(
                f"\n警告: 最大深度({max_depth})に到達しました。"
                "ツリーが切り捨てられている可能性があります。",
                file=sys.stderr,
            )
            print(
                f"ヒント: --depth オプションで深度を増やすことを検討してください。",
                file=sys.stderr,
            )

    def print_reverse_tree(
        self,
        target_method: str,
        max_depth: int = 50,
        show_class: bool = False,
        follow_overrides: bool = True,
        verbose: bool = False,
        use_tab: bool = False,
        short_mode: bool = False,
    ):
        """呼び出し先からのツリー（誰がこのメソッドを呼んでいるか）を表示

        Args:
            target_method: 対象メソッド
            max_depth: 最大深度
            show_class: クラス情報を表示するか
            follow_overrides: オーバーライド元/インターフェースメソッドも追跡するか
            verbose: 詳細表示（Javadocを表示）
            use_tab: Trueの場合、ハードタブでインデントし、プレフィックスを省略
            short_mode: Trueの場合、クラス名からパッケージ名を省いて表示
        """
        print(f"\n{'=' * 80}")
        print(f"逆引きツリー (対象: {target_method})")
        print(f"{'=' * 80}\n")

        visited: set[str] = set()
        final_endpoints: set[str] = set()  # 最終到達点のメソッドを収集
        # 最大深度到達フラグを初期化
        max_depth_reached: List[bool] = [False]

        self._print_reverse_tree_recursive(
            target_method,
            0,
            max_depth,
            visited,
            show_class,
            follow_overrides,
            final_endpoints,
            verbose,
            use_tab,
            short_mode,
            max_depth_reached,
        )

        # 最終到達点のメソッド一覧を表示
        if final_endpoints:
            print(f"\n{'=' * 80}")
            print(f"最終到達点のメソッド一覧 (最上位の呼び元メソッド)")
            print(f"{'=' * 80}\n")
            for endpoint in sorted(final_endpoints):
                if short_mode:
                    endpoint = self._shorten_method_signature(endpoint)
                print(f"  {endpoint}")
            print()

        # 最大深度に到達した場合の警告を出力
        if max_depth_reached[0]:
            print(
                f"\n警告: 最大深度({max_depth})に到達しました。"
                "ツリーが切り捨てられている可能性があります。",
                file=sys.stderr,
            )
            print(
                f"ヒント: --depth オプションで深度を増やすことを検討してください。",
                file=sys.stderr,
            )

    def _print_tree_recursive(
        self,
        method: str,
        depth: int,
        max_depth: int,
        visited: Set[str],
        show_class: bool,
        show_sql: bool,
        is_forward: bool,
        follow_implementations: bool = True,
        verbose: bool = False,
        use_tab: bool = False,
        short_mode: bool = False,
        accumulated_instances: Optional[Set[str]] = None,  # 累積されたインスタンス情報
        max_depth_reached: Optional[List[bool]] = None,  # 最大深度到達フラグ
    ):
        """ツリーを再帰的に表示

        Args:
            accumulated_instances: 呼び出しツリーの上位から累積された生成インスタンス情報
            max_depth_reached: 最大深度到達フラグ（[False]のリストで渡し、到達時に[True]に更新）
        """
        if depth > max_depth:
            # 最大深度に到達した場合、フラグをセット
            if max_depth_reached is not None:
                max_depth_reached[0] = True
            return

        # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
        if not self.exclusion_manager.should_include(method):
            return

        # 循環参照チェック
        if method in visited:
            self._print_node(
                method,
                depth,
                show_class,
                show_sql,
                is_circular=True,
                verbose=verbose,
                use_tab=use_tab,
                short_mode=short_mode,
            )
            return

        visited.add(method)
        self._print_node(
            method,
            depth,
            show_class,
            show_sql,
            verbose=verbose,
            use_tab=use_tab,
            short_mode=short_mode,
        )

        # 現在のメソッドで生成されるインスタンスを収集し、累積に追加
        current_instances = self._collect_created_instances(method)
        if accumulated_instances is None:
            accumulated_instances = current_instances
        else:
            accumulated_instances.update(current_instances)

        # Eモード: 除外対象の場合、配下の展開を停止
        if self.exclusion_manager.should_exclude_children(method):
            indent = "\t" * (depth + 1) if use_tab else "    " * (depth + 1)
            print(f"{indent}〓[配下の呼び出しを除外]")
            return

        # 子ノードを表示
        if is_forward:
            callees = self.forward_calls.get(method, [])
            for callee_info in callees:
                callee = callee_info["method"]

                indent = "\t" * (depth + 1) if use_tab else "    " * (depth + 1)

                # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                if not self.exclusion_manager.should_include(callee):
                    continue

                # 親クラスメソッドの情報を表示
                if callee_info["is_parent_method"] == "Yes":
                    print(f"{indent}〓↓ [親クラスメソッド]")

                # 呼び出し先を再帰的に表示
                self._print_tree_recursive(
                    callee,
                    depth + 1,
                    max_depth,
                    visited.copy(),
                    show_class,
                    show_sql,
                    is_forward,
                    follow_implementations,
                    verbose,
                    use_tab,
                    short_mode,
                    accumulated_instances,
                    max_depth_reached,
                )

                # 実装クラス候補の情報を表示
                if callee_info["implementations"]:
                    implementations = [
                        impl.strip()
                        for impl in callee_info["implementations"].split(",")
                        if impl.strip()
                    ]

                    annotations = []
                    for impl_class_info in implementations:
                        # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                        impl_class = impl_class_info.split(" ")[0]
                        if not self.exclusion_manager.should_include(impl_class):
                            continue
                        if self.debug_mode:
                            annotations.append(f"実装: {impl_class_info}")

                    for annotation in annotations:
                        indent = "\t" * (depth + 1) if use_tab else "    " * (depth + 1)
                        print(f"{indent}〓^ [{annotation}]")

                    # 実装クラス候補がある場合、それらも追跡
                    if follow_implementations:
                        # implementationsの各要素は、「<クラス名> + " [<追加情報>]"」の形式かもしれないので、クラス名だけ抽出
                        impl_classes = [impl.split(" ")[0] for impl in implementations]

                        # Eモード: 除外対象の場合、実装クラスへの展開を停止
                        if self.exclusion_manager.should_exclude_children(callee):
                            indent = (
                                "\t" * (depth + 1) if use_tab else "    " * (depth + 1)
                            )
                            print(f"{indent}〓[実装クラスへの展開を除外]")
                            continue

                        # 累積されたインスタンス情報に基づいてフィルタリング
                        if accumulated_instances:
                            filtered_impl_classes = (
                                self._filter_implementations_by_accumulated_instances(
                                    accumulated_instances, impl_classes
                                )
                            )
                        else:
                            filtered_impl_classes = impl_classes

                        for impl_class in filtered_impl_classes:
                            # 実装クラスの対応するメソッドを探す
                            impl_method = self._find_implementation_method(
                                callee, impl_class
                            )
                            if impl_method:
                                # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                                if not self.exclusion_manager.should_include(
                                    impl_method
                                ):
                                    continue

                                indent = (
                                    "\t" * (depth + 1)
                                    if use_tab
                                    else "    " * (depth + 1)
                                )
                                print(f"{indent}〓> [実装クラスへの展開: {impl_class}]")

                                self._print_tree_recursive(
                                    impl_method,
                                    depth + 1,
                                    max_depth,
                                    visited.copy(),
                                    show_class,
                                    show_sql,
                                    is_forward,
                                    follow_implementations,
                                    verbose,
                                    use_tab,
                                    short_mode,
                                    accumulated_instances,
                                    max_depth_reached,
                                )
        else:
            callers = self.reverse_calls.get(method, [])
            for caller in callers:
                self._print_tree_recursive(
                    caller,
                    depth + 1,
                    max_depth,
                    visited.copy(),
                    show_class,
                    False,
                    is_forward,
                    follow_implementations,
                    verbose,
                    use_tab,
                    short_mode,
                    accumulated_instances,
                    max_depth_reached,
                )

    def _print_reverse_tree_recursive(
        self,
        method: str,
        depth: int,
        max_depth: int,
        visited: Set[str],
        show_class: bool,
        follow_overrides: bool,
        final_endpoints: Optional[Set[str]] = None,
        verbose: bool = False,
        use_tab: bool = False,
        short_mode: bool = False,
        max_depth_reached: Optional[List[bool]] = None,  # 最大深度到達フラグ
    ):
        """逆引きツリーを再帰的に表示"""
        if depth > max_depth:
            # 最大深度に到達した場合、フラグをセット
            if max_depth_reached is not None:
                max_depth_reached[0] = True
            return

        # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
        if not self.exclusion_manager.should_include(method):
            return

        # 循環参照チェック
        if method in visited:
            self._print_node(
                method,
                depth,
                show_class,
                False,
                is_circular=True,
                verbose=verbose,
                use_tab=use_tab,
                short_mode=short_mode,
            )
            return

        visited.add(method)
        self._print_node(
            method,
            depth,
            show_class,
            False,
            verbose=verbose,
            use_tab=use_tab,
            short_mode=short_mode,
        )

        callers = self.reverse_calls.get(method, [])

        # 呼び出し元がない場合、オーバーライド元/インターフェースメソッドを探す
        if not callers and follow_overrides:
            parent_methods = self._find_parent_methods(method)
            if parent_methods:
                indent = "\t" * depth if use_tab else "    " * depth
                print(f"{indent}〓> [オーバーライド元/インターフェースメソッドを展開]")
                for parent_method in parent_methods:
                    self._print_reverse_tree_recursive(
                        parent_method,
                        depth,
                        max_depth,
                        visited.copy(),
                        show_class,
                        follow_overrides,
                        final_endpoints,
                        verbose,
                        use_tab,
                        short_mode,
                        max_depth_reached,
                    )
            else:
                # オーバーライド元もない場合は最終到達点
                if final_endpoints is not None:
                    final_endpoints.add(method)
        elif not callers:
            # 呼び出し元がない場合は最終到達点
            if final_endpoints is not None:
                final_endpoints.add(method)
        else:
            # 通常の呼び出し元を表示
            for caller in callers:
                self._print_reverse_tree_recursive(
                    caller,
                    depth + 1,
                    max_depth,
                    visited.copy(),
                    show_class,
                    follow_overrides,
                    final_endpoints,
                    verbose,
                    use_tab,
                    short_mode,
                    max_depth_reached,
                )

    def _print_node(
        self,
        method: str,
        depth: int,
        show_class: bool,
        show_sql: bool,
        is_circular: bool = False,
        verbose: bool = False,
        use_tab: bool = False,
        short_mode: bool = False,
    ):
        """ノード情報を表示"""
        # use_tabがTrueの場合、ハードタブでインデントし、プレフィックスを省略
        if use_tab:
            indent = "\t" * depth
            prefix = ""
        else:
            indent = "    " * depth
            prefix = "|-- " if depth > 0 else ""

        info = self.method_info.get(method, {})

        # 表示するメソッド名を決定（short_modeの場合、パッケージ名を省略）
        display_method = (
            self._shorten_method_signature(method) if short_mode else method
        )

        # メソッド名を表示
        display = f"{indent}{prefix}{display_method}"
        if is_circular:
            display += " [循環参照]"

        # verboseモードの場合、Javadocをタブ区切りで追加
        if verbose:
            javadoc = info.get("javadoc", "")
            if javadoc:
                display += f"    〓{javadoc}"

        print(display)

        # クラス情報を表示
        if show_class and info.get("class"):
            class_name = info.get("class", "")
            sub_indent = "    "
            print(f"{indent}{sub_indent}〓クラス: {class_name}")
            if info.get("parent"):
                parent_class = info.get("parent", "")
                print(f"{indent}{sub_indent}〓親クラス: {parent_class}")

        # SQL情報を表示（全文表示）
        if show_sql and info.get("sql"):
            sql_text = info.get("sql", "") or ""
            sub_indent = "    "
            print(f"{indent}{sub_indent}〓SQL: {sql_text}")

    def _shorten_method_signature(self, method: str) -> str:
        """メソッドシグネチャからパッケージ名を省いて返す

        Args:
            method: メソッドシグネチャ (例: "com.example.MyClass#myMethod(com.example.Arg)")

        Returns:
            パッケージ名を省いたメソッドシグネチャ (例: "MyClass#myMethod(Arg)")
        """
        if "#" not in method:
            # メソッドシグネチャでない場合はそのまま返す
            return method

        # クラス名#メソッド名(引数) の形式を解析
        hash_pos = method.find("#")
        class_part = method[:hash_pos]
        method_part = method[hash_pos + 1 :]

        # クラス名からパッケージ名を省略
        short_class = class_part.split(".")[-1] if "." in class_part else class_part

        # メソッド名の引数部分もパッケージ名を省略
        paren_pos = method_part.find("(")
        if paren_pos >= 0:
            method_name = method_part[:paren_pos]
            args_part = method_part[paren_pos + 1 : -1]  # () 内の部分
            if args_part:
                # 引数をカンマで分割し、各引数のパッケージ名を省略
                short_args = []
                for arg in args_part.split(","):
                    arg = arg.strip()
                    # 配列の場合 (e.g., "String[]", "com.example.Type[]")
                    array_suffix = ""
                    if arg.endswith("[]"):
                        array_suffix = "[]"
                        arg = arg[:-2]
                    # パッケージ名を省略
                    short_arg = arg.split(".")[-1] if "." in arg else arg
                    short_args.append(short_arg + array_suffix)
                short_method_part = f"{method_name}({', '.join(short_args)})"
            else:
                short_method_part = f"{method_name}()"
        else:
            short_method_part = method_part

        return f"{short_class}#{short_method_part}"

    def _find_parent_methods(self, method: str) -> List[str]:
        """メソッドのオーバーライド元/インターフェースメソッドを探す

        Args:
            method: 対象メソッドのシグネチャ

        Returns:
            親メソッドのシグネチャのリスト
        """
        if "#" not in method:
            return []

        info = self.method_info.get(method, {})
        method_part = method.split("#", 1)[1]  # メソッド名+引数部分
        parent_classes_str = info.get("parent", "")

        if not parent_classes_str:
            return []

        # 親クラスをカンマで分割
        parent_classes = [p.strip() for p in parent_classes_str.split(",") if p.strip()]

        parent_methods = []
        for parent_class in parent_classes:
            # 親クラスの同じシグネチャのメソッドを探す
            for method_sig, method_info in self.method_info.items():
                if method_info.get("class") == parent_class:
                    if "#" in method_sig:
                        sig_method_part = method_sig.split("#", 1)[1]
                        if sig_method_part == method_part:
                            parent_methods.append(method_sig)

        return parent_methods

    def _find_implementation_method(
        self, abstract_method: str, impl_class: str
    ) -> Optional[str]:
        """抽象メソッドに対応する実装クラスのメソッドを探す

        Args:
            abstract_method: 抽象メソッドのシグネチャ
            impl_class: 実装クラス名

        Returns:
            実装メソッドのシグネチャ（見つからない場合はNone）
        """
        # メソッドシグネチャからメソッド名と引数を抽出
        # 例: "com.example.Interface#method(String, int)" -> "method(String, int)"
        if "#" not in abstract_method:
            return None

        method_part = abstract_method.split("#", 1)[1]

        # 実装クラスの同じシグネチャのメソッドを探す
        # 1. 直接の実装を探す
        for method_sig, info in self.method_info.items():
            if info.get("class") == impl_class:
                if "#" in method_sig:
                    sig_method_part = method_sig.split("#", 1)[1]
                    if sig_method_part == method_part:
                        return method_sig

        # 2. 親クラスを辿って実装を探す
        current_class = impl_class
        visited_classes = set()

        while current_class:
            if current_class in visited_classes:
                break
            visited_classes.add(current_class)

            parents = self.class_info.get(current_class, [])
            if not parents:
                break

            # 親クラス（インターフェース含む）の中から実装を探す
            # 優先順位: クラス > インターフェース だが、ここでは見つかった順
            # Javaの単一継承を考えると、親クラスは1つ（+インターフェース複数）
            # ここでは単純に見つかったものを返す
            found_in_parent = False
            for parent in parents:
                # インターフェースはスキップ
                if parent in self.interface_data:
                    continue

                # 親クラスでメソッドを探す
                for method_sig, info in self.method_info.items():
                    if info.get("class") == parent:
                        if "#" in method_sig:
                            sig_method_part = method_sig.split("#", 1)[1]
                            if sig_method_part == method_part:
                                return method_sig

                # 見つからなければ、次のループのために親クラスを更新したいが、
                # 複数の親（インターフェース）があるため、幅優先探索すべきか？
                # ここでは単純化して、最初の親（おそらくスーパークラス）を優先して探索する
                # ただし、parentsリストの順序は保証されていないかもしれない
                pass

            # 幅優先探索で次の階層へ
            # parents の中から次の current_class を選ぶ必要があるが、
            # ここでは簡易的に、parents のすべてを次の探索候補にするBFSにする

            # 上のループで return していないので、ここに来たということは
            # 直近の親には実装がない。
            # 次の階層（親の親）を探す

            # BFS queue logic implementation
            # Since we are inside a while loop designed for single path, let's refactor to BFS queue
            break  # Break inner loop to switch to BFS structure below

        # BFSで親クラスを探索
        queue = list(self.class_info.get(impl_class, []))
        visited_classes = {impl_class}

        while queue:
            current_parent = queue.pop(0)
            if current_parent in visited_classes:
                continue
            visited_classes.add(current_parent)

            # インターフェースはスキップ
            if current_parent in self.interface_data:
                continue

            # この親クラスにメソッドがあるか確認
            for method_sig, info in self.method_info.items():
                if info.get("class") == current_parent:
                    if "#" in method_sig:
                        sig_method_part = method_sig.split("#", 1)[1]
                        if sig_method_part == method_part:
                            return method_sig

            # 次の親を追加
            queue.extend(self.class_info.get(current_parent, []))

        return None

    def _filter_implementations_by_accumulated_instances(
        self,
        accumulated_instances: Set[str],
        implementations: List[str],
    ) -> List[str]:
        """累積されたインスタンス情報を使用して実装クラス候補をフィルタリング

        Args:
            accumulated_instances: 呼び出しツリーの上位から累積された生成インスタンス
            implementations: 実装クラス候補のリスト

        Returns:
            フィルタリングされた実装クラスのリスト
        """
        if not implementations or not accumulated_instances:
            return implementations

        # 実装クラス候補が累積インスタンスに含まれるかチェック
        filtered = [impl for impl in implementations if impl in accumulated_instances]

        if filtered:
            return filtered  # マッチするものがあればそれのみ返す

        return implementations  # マッチしなければ全候補を返す

    def _collect_created_instances(self, method: str) -> Set[str]:
        """メソッドおよびそのクラスで生成されるインスタンスを収集

        Args:
            method: メソッドシグネチャ

        Returns:
            生成されるインスタンスのクラス名セット
        """
        created_instances: Set[str] = set()

        # メソッド内で生成されたインスタンス
        method_info = self.method_info.get(method, {})
        created_instances.update(method_info.get("createdInstances", []))

        # クラスのフィールド初期化で生成されたインスタンス
        method_class = method_info.get("class", "")
        if method_class:
            class_data = self.class_data.get(method_class, {})
            for init in class_data.get("fieldInitializers", []):
                initialized_class = init.get("initializedClass", "")
                if initialized_class:
                    created_instances.add(initialized_class)

        # デバッグモードの場合、収集したインスタンス情報を出力
        if self.debug_mode and created_instances:
            print(
                f"[DEBUG] {method} で収集したインスタンス: {', '.join(sorted(created_instances))}"
            )

        return created_instances

    def export_tree_to_file(
        self,
        root_method: str,
        output_file: str,
        max_depth: int = 10,
        format: str = "text",
        follow_implementations: bool = True,
    ):
        """ツリーをファイルにエクスポート"""
        if format == "text":
            self._export_text_tree(
                root_method, output_file, max_depth, follow_implementations
            )
        elif format == "markdown":
            self._export_markdown_tree(
                root_method, output_file, max_depth, follow_implementations
            )
        elif format == "html":
            self._export_html_tree(
                root_method, output_file, max_depth, follow_implementations
            )

    def _export_text_tree(
        self,
        root_method: str,
        output_file: str,
        max_depth: int,
        follow_implementations: bool,
    ):
        """テキスト形式でエクスポート"""
        with open(output_file, "w", encoding="utf-8") as f:
            original_stdout = sys.stdout
            sys.stdout = f
            self.print_forward_tree(
                root_method, max_depth, follow_implementations=follow_implementations
            )
            sys.stdout = original_stdout
        print(f"ツリーを {output_file} にエクスポートしました")

    def _export_markdown_tree(
        self,
        root_method: str,
        output_file: str,
        max_depth: int,
        follow_implementations: bool,
    ):
        """Markdown形式でエクスポート"""
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(f"# 呼び出しツリー\n\n")
            f.write(f"**起点メソッド:** `{root_method}`\n\n")
            f.write("```\n")

            original_stdout = sys.stdout
            sys.stdout = f
            self.print_forward_tree(
                root_method,
                max_depth,
                show_class=False,
                show_sql=False,
                follow_implementations=follow_implementations,
            )
            sys.stdout = original_stdout

            f.write("```\n")
        print(f"ツリーを {output_file} にエクスポートしました")

    def _export_html_tree(
        self,
        root_method: str,
        output_file: str,
        max_depth: int,
        follow_implementations: bool,
    ):
        """HTML形式でエクスポート（インタラクティブなツリー）"""
        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>呼び出しツリー - {root_method}</title>
    <style>
        body {{ font-family: 'Courier New', monospace; padding: 20px; }}
        .tree {{ list-style-type: none; padding-left: 20px; }}
        .tree li {{ margin: 5px 0; }}
        .method {{ color: #0066cc; cursor: pointer; }}
        .method:hover {{ text-decoration: underline; }}
        .class-info {{ color: #666; font-size: 0.9em; margin-left: 20px; }}
        .sql {{ color: #008800; font-size: 0.9em; margin-left: 20px; }}
        .circular {{ color: #cc0000; }}
        .parent-method {{ color: #ff6600; }}
        .implementation {{ color: #9966cc; font-style: italic; }}
    </style>
</head>
<body>
    <h1>呼び出しツリー</h1>
    <p><strong>起点メソッド:</strong> {root_method}</p>
    <ul class="tree">
"""

        visited = set()
        html += self._generate_html_tree(
            root_method, 0, max_depth, visited, follow_implementations
        )

        html += """
    </ul>
</body>
</html>
"""

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"ツリーを {output_file} にエクスポートしました")

    def _generate_html_tree(
        self,
        method: str,
        depth: int,
        max_depth: int,
        visited: Set[str],
        follow_implementations: bool,
    ) -> str:
        """HTML形式のツリーを生成"""
        if depth > max_depth:
            return ""

        # Iモード: 除外対象の場合、ノード自体をスキップ
        if not self.exclusion_manager.should_include(method):
            return ""

        html = ""
        info = self.method_info.get(method, {})

        if method in visited:
            html += f'<li><span class="method circular">{method} [循環参照]</span></li>'
            return html

        visited.add(method)

        html += f'<li><span class="method">{method}</span>'

        if info.get("class"):
            html += f'<div class="class-info">クラス: {info["class"]}</div>'

        # Eモード: 配下の展開を停止
        if self.exclusion_manager.should_exclude_children(method):
            html += '<div class="class-info">[配下の呼び出しを除外]</div>'
            html += "</li>"
            return html

        callees = self.forward_calls.get(method, [])
        if callees:
            html += '<ul class="tree">'
            for callee_info in callees:

                # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                if not self.exclusion_manager.should_include(callee_info["method"]):
                    continue

                html += self._generate_html_tree(
                    callee_info["method"],
                    depth + 1,
                    max_depth,
                    visited.copy(),
                    follow_implementations,
                )

                # 実装クラス候補がある場合
                if follow_implementations and callee_info["implementations"]:
                    implementations = [
                        impl.strip()
                        for impl in callee_info["implementations"].split(",")
                        if impl.strip()
                    ]
                    # implementationsの各要素は、「<クラス名> + " [<追加情報>]"」の形式かもしれないので、クラス名だけ抽出
                    implementations = [impl.split(" ")[0] for impl in implementations]

                    for impl_class in implementations:
                        impl_method = self._find_implementation_method(
                            callee_info["method"], impl_class
                        )
                        if impl_method:
                            # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                            if not self.exclusion_manager.should_include(impl_method):
                                continue

                            html += f'<li><span class="implementation">→ 実装: {impl_class}</span>'
                            html += self._generate_html_tree(
                                impl_method,
                                depth + 2,
                                max_depth,
                                visited.copy(),
                                follow_implementations,
                            )
                            html += "</li>"

            html += "</ul>"

        html += "</li>"
        return html

    def list_entry_points(self, min_calls: int = 1, strict: bool = True):
        """エントリーポイント候補をリストアップ

        Args:
            min_calls: 最小呼び出し数
            strict: True の場合、アノテーションやmainメソッドなど厳密に判定（デフォルト）
        """
        print(f"\n{'=' * 80}")
        if strict:
            print("エントリーポイント候補（厳密モード）")
        else:
            print(f"エントリーポイント候補 (呼び出し先が{min_calls}個以上)")
        print(f"{'=' * 80}\n")

        # すべての呼び出し先を収集
        all_callees = set()
        for callees in self.forward_calls.values():
            for callee_info in callees:
                all_callees.add(callee_info["method"])

        entry_points = []

        for method, info in self.method_info.items():
            # 他から呼ばれていないメソッドのみ
            if method in all_callees:
                continue

            # インターフェースの場合は除外
            type = info.get("class", "")
            if self.interface_data.get(type, ""):
                continue

            call_count = len(self.forward_calls.get(method, []))

            # 厳密モードの場合
            if strict:
                if info.get("is_entry_point"):
                    entry_type = self._determine_entry_type(method, info)
                    entry_points.append(
                        (
                            method,
                            call_count,
                            info.get("class", ""),
                            entry_type,
                            info.get("annotations", ""),
                            info.get("visibility", ""),
                            info.get("javadoc", ""),
                        )
                    )
            else:
                # 非厳密モードの場合は呼び出し数で判定
                if call_count >= min_calls:
                    entry_type = self._determine_entry_type(method, info)
                    entry_points.append(
                        (
                            method,
                            call_count,
                            info.get("class", ""),
                            entry_type,
                            info.get("annotations", ""),
                            info.get("visibility", ""),
                            info.get("javadoc", ""),
                        )
                    )

        # エントリータイプとメソッド名でソート
        entry_points.sort(key=lambda x: (self._entry_priority(x[3]), x[0]))

        # 結果を表示
        if not entry_points:
            print("エントリーポイント候補が見つかりませんでした", file=sys.stderr)
            return

        for i, (
            method,
            call_count,
            class_name,
            entry_type,
            annotations,
            visibility,
            javadoc,
        ) in enumerate(entry_points, 1):
            print(f"{i}. {method}")
            print(f"   クラス: {class_name}")
            print(f"   種別: {entry_type}")
            print(f"   Javadoc: {javadoc}")
            print(f"   可視性: {visibility}")
            if annotations:
                print(f"   メソッドアノテーション: {annotations}")

            # クラスアノテーションも表示（親クラス・インターフェース含む）
            info = self.method_info.get(method, {})
            all_class_annotations = self._get_all_class_annotation_raws(class_name)
            if all_class_annotations:
                print(f"   クラスアノテーション: {', '.join(all_class_annotations)}")

            # HTTP / SOAP の場合、アノテーション等からエンドポイントの path を抽出して表示
            if entry_type and ("HTTP Endpoint" in entry_type or "SOAP" in entry_type):
                endpoint_path = self._extract_endpoint_path(info, class_name)
                if endpoint_path:
                    print(f"   エンドポイント: {endpoint_path}")

            print(f"   呼び出し数: {call_count}")
            print()

    def list_entry_points_tsv(self, min_calls: int = 1, strict: bool = True) -> None:
        """エントリーポイント候補をTSV形式で出力

        Args:
            min_calls: 最小呼び出し数
            strict: True の場合、アノテーションやmainメソッドなど厳密に判定（デフォルト）
        """
        #  clipでコピーした結果をExcelに貼り付けられるにはShift_JISで出力する
        sys.stdout.reconfigure(encoding=self.output_tsv_encoding)

        # すべての呼び出し先を収集
        all_callees = set()
        for callees in self.forward_calls.values():
            for callee_info in callees:
                all_callees.add(callee_info["method"])

        entry_points = []

        for method, info in self.method_info.items():
            # 他から呼ばれていないメソッドのみ
            if method in all_callees:
                continue

            # インターフェースの場合は除外
            type = info.get("class", "")
            if self.interface_data.get(type, ""):
                continue

            call_count = len(self.forward_calls.get(method, []))

            # 厳密モードの場合
            if strict:
                if info.get("is_entry_point"):
                    entry_type = self._determine_entry_type(method, info)
                    entry_points.append(
                        (
                            method,
                            call_count,
                            info.get("class", ""),
                            entry_type,
                            info.get("annotations", ""),
                            info.get("visibility", ""),
                            info.get("javadoc", ""),
                        )
                    )
            else:
                # 非厳密モードの場合は呼び出し数で判定
                if call_count >= min_calls:
                    entry_type = self._determine_entry_type(method, info)
                    entry_points.append(
                        (
                            method,
                            call_count,
                            info.get("class", ""),
                            entry_type,
                            info.get("annotations", ""),
                            info.get("visibility", ""),
                            info.get("javadoc", ""),
                        )
                    )

        # エントリータイプとメソッド名でソート
        entry_points.sort(key=lambda x: (self._entry_priority(x[3]), x[0]))

        # TSVヘッダーを出力
        print(
            "メソッド\tパッケージ名\tクラス名\tメソッド名\tエンドポイント\tメソッドjavadoc\tクラスjavadoc\t種別\tメソッドアノテーション\tクラスアノテーション"
        )

        # 結果をTSV形式で出力
        for (
            method,
            call_count,
            class_name,
            entry_type,
            annotations,
            visibility,
            javadoc,
        ) in entry_points:
            # パッケージ名とクラス名（パッケージ除く）を分離
            package_name = ""
            class_name_only = class_name
            if "." in class_name:
                parts = class_name.rsplit(".", 1)
                package_name = parts[0]
                class_name_only = parts[1]

            # メソッド名（クラスや引数を含めないメソッド名のみ）を抽出
            method_name_only = ""
            if "#" in method:
                # クラス#メソッド(引数) の形式からメソッド名のみを抽出
                method_part = method.split("#", 1)[1]
                # 引数部分を除去
                if "(" in method_part:
                    method_name_only = method_part.split("(", 1)[0]
                else:
                    method_name_only = method_part
            else:
                # #がない場合はそのまま
                if "(" in method:
                    method_name_only = method.split("(", 1)[0]
                else:
                    method_name_only = method

            # エンドポイントを抽出
            info = self.method_info.get(method, {})
            endpoint_path = ""
            if entry_type and ("HTTP Endpoint" in entry_type or "SOAP" in entry_type):
                endpoint_path = self._extract_endpoint_path(info, class_name)

            # クラスアノテーションとクラスJavadocを取得（親クラス・インターフェース含む）
            all_class_annotations = self._get_all_class_annotation_raws(class_name)
            class_annotations_str = ", ".join(all_class_annotations)
            class_javadoc = self._get_class_javadoc(class_name)

            # TSV行を出力
            print(
                f"{method}\t{package_name}\t{class_name_only}\t{method_name_only}\t{endpoint_path}\t{javadoc}\t{class_javadoc}\t{entry_type}\t{annotations}\t{class_annotations_str}"
            )

    def _extract_endpoint_path(self, info: dict, class_name: str = "") -> str:
        """アノテーションやクラスアノテーションからエンドポイントの path を抽出する

        親クラス・インターフェースのアノテーションも考慮し、
        クラスレベルのパス + メソッドレベルのパスを結合して返す。

        戻り値: 見つかれば path（例: /api/foo や https://... など）、見つからなければ空文字
        """
        if not info:
            return ""

        method_annotations = str(info.get("annotations", ""))

        # クラスレベルのアノテーション（親クラス・インターフェース含む、フル形式）を取得
        all_class_annotation_raws = (
            self._get_all_class_annotation_raws(class_name) if class_name else []
        )
        class_annotations = " ".join(all_class_annotation_raws)

        # パス抽出パターン
        path_patterns = [
            r"\w*Mapping\(\s*path\s*=\s*[\"']([^\"']+)[\"']",  # path = "/x"
            r"\w*Mapping\(\s*value\s*=\s*[\"']([^\"']+)[\"']",  # value = "/x"
            r"\w*Mapping\(\s*[\"']([^\"']+)[\"']",  # GetMapping("/x"), RequestMapping("/x") 等
            r"Path\(\s*[\"']([^\"']+)[\"']",  # JAX-RS @Path
        ]

        # クラスレベルの基本パスを抽出（@RequestMapping等から）
        base_path = ""
        for pattern in path_patterns:
            m = re.search(pattern, class_annotations)
            if m:
                base_path = m.group(1)
                break

        # メソッドレベルのパスを抽出
        method_path = ""
        for pattern in path_patterns:
            m = re.search(pattern, method_annotations)
            if m:
                method_path = m.group(1)
                break

        # パスの結合
        if base_path and method_path:
            # 両方存在する場合は結合
            if base_path.endswith("/"):
                base_path = base_path[:-1]
            if not method_path.startswith("/"):
                method_path = "/" + method_path
            return base_path + method_path
        elif base_path:
            return base_path
        elif method_path:
            return method_path

        # WebLogic + JAX-WS（SOAP）の場合を考慮し、WebLogic特有アノテーション @WLHttpTransportのcontextPath、serviceUriの値からパスを抽出
        # さらに、メソッドレベルの@WebMethodのoperationNameの値を結合して、SOAPのエンドポイントを生成する
        # 例：contextPath=/foo, serviceUri=/bar, operationName=fooBar -> /foo/bar : operationName=fooBar
        context_path = ""
        service_uri = ""
        operation_name = ""
        m = re.search(r"contextPath\s*=\s*[\"']([^\"']+)[\"']", class_annotations)
        if m:
            context_path = m.group(1)
        m = re.search(r"serviceUri\s*=\s*[\"']([^\"']+)[\"']", class_annotations)
        if m:
            service_uri = m.group(1)
        m = re.search(r"operationName\s*=\s*[\"']([^\"']+)[\"']", method_annotations)
        if m:
            operation_name = m.group(1)
        if context_path or service_uri or operation_name:
            return context_path + service_uri + " : operationName=" + operation_name

        # # 汎用パターンでの検索（上記で見つからない場合）
        # full_text = method_annotations + " " + class_annotations
        # general_patterns = [
        #     r"[\"'](/[^\"']+)[\"']",  # 汎用: 引用された /... パス
        #     r"[\"'](https?://[^\"']+)[\"']",  # フルURL
        # ]
        # for pattern in general_patterns:
        #     m = re.search(pattern, full_text)
        #     if m:
        #         return m.group(1)

        # # SOAP系: class アノテーションに serviceName や targetNamespace があれば返す
        # m2 = re.search(
        #     r"(?:serviceName|targetNamespace)\s*=\s*[\"']([^\"']+)[\"']", full_text
        # )
        # if m2:
        #     return m2.group(1)

        return ""

    def _determine_entry_type(self, method: str, info: dict) -> str:
        """エントリーポイントの種別を判定"""
        candidates = []

        # 1. 解析時に判定されたエントリータイプ
        entry_type = info.get("entry_type", "")
        if entry_type:
            type_map = {
                "Main": "Main Method",
                "Test": "Test Method",
                "HTTP": "HTTP Endpoint",
                "SOAP": "SOAP Endpoint",
                "Scheduled": "Scheduled Job",
                "Event": "Event Listener",
                "Lifecycle": "Lifecycle Method",
                "Servlet": "Servlet",
                "SpringBoot": "Spring Boot Runner",
                "Thread": "Runnable/Callable",
                "Bean": "Bean Factory",
            }
            candidates.append(type_map.get(entry_type, entry_type))

        # 2. アノテーションから判定
        entry_type = self._check_entry_type_from_info(info)
        if entry_type:
            candidates.append(entry_type)

        # 3. 親メソッド（インターフェース/スーパークラス）のアノテーションを確認
        parent_methods = self._find_parent_methods(method)
        for parent_method in parent_methods:
            parent_info = self.method_info.get(parent_method, {})
            if parent_info:
                entry_type = self._check_entry_type_from_info(parent_info)
                if entry_type:
                    candidates.append(entry_type)

        if not candidates:
            return ""

        # 優先順位でソート（数値が小さい方が優先）
        candidates.sort(key=lambda x: self._entry_priority(x))
        return candidates[0]

    def _check_entry_type_from_info(self, info: dict) -> str:
        """メソッド情報からエントリータイプを判定（ヘルパー）"""
        annotations = info.get("annotations", "")
        class_annotations = info.get("class_annotations", "")

        # main メソッド
        if info.get("is_static") and "main" in info.get("class", "").lower():
            return "Main Method"

        # テストメソッド
        if any(
            test in annotations
            for test in [
                "Test",
                "TestTemplate",
                "ParameterizedTest",
                "RepeatedTest",
                "TestFactory",
            ]
        ):
            return "Test Method"

        # Spring Controller
        if any(
            ctrl in annotations
            for ctrl in [
                "RequestMapping",
                "GetMapping",
                "PostMapping",
                "PutMapping",
                "DeleteMapping",
                "PatchMapping",
            ]
        ):
            return "HTTP Endpoint (Spring)"

        # クラスレベルのController
        if any(ctrl in class_annotations for ctrl in ["Controller", "RestController"]):
            return "HTTP Endpoint (Spring)"

        # JAX-RS REST API
        if any(
            jax in annotations
            for jax in ["Path", "GET", "POST", "PUT", "DELETE", "PATCH"]
        ):
            return "HTTP Endpoint (JAX-RS)"

        # SOAP Webサービス
        if "WebMethod" in annotations:
            return "SOAP Endpoint (JAX-WS)"

        if any(ws in class_annotations for ws in ["WebService", "WebServiceProvider"]):
            return "SOAP Endpoint (JAX-WS)"

        # Scheduled Job
        if any(sched in annotations for sched in ["Scheduled", "Schedules", "Async"]):
            return "Scheduled Job"

        # Event Listener
        if any(
            evt in annotations
            for evt in [
                "EventListener",
                "TransactionalEventListener",
                "JmsListener",
                "RabbitListener",
                "KafkaListener",
                "StreamListener",
                "MessageMapping",
                "SubscribeMapping",
            ]
        ):
            return "Event Listener"

        # Lifecycle
        if any(
            lc in annotations
            for lc in [
                "PostConstruct",
                "PreDestroy",
                "BeforeAll",
                "AfterAll",
                "BeforeEach",
                "AfterEach",
                "Before",
                "After",
                "BeforeClass",
                "AfterClass",
            ]
        ):
            return "Lifecycle Method"

        # Bean Factory Method
        if "Bean" in annotations:
            return "Bean Factory"

        # Servlet
        if "WebServlet" in annotations or "WebServlet" in class_annotations:
            return "Servlet"

        # その他のpublicメソッド
        if info.get("visibility") == "public":
            return "Public Method"

        return "Unknown"

    def _entry_priority(self, entry_type: str) -> int:
        """エントリータイプの優先順位（小さいほど優先度が高い）"""
        priority_map = {
            "Main Method": 1,
            "HTTP Endpoint (Spring)": 2,
            "HTTP Endpoint (JAX-RS)": 3,
            "SOAP Endpoint (JAX-WS)": 4,
            "Servlet": 5,
            "Spring Boot Runner": 6,
            "Scheduled Job": 7,
            "Event Listener": 8,
            "Runnable/Callable": 9,
            "Lifecycle Method": 10,
            "Test Method": 11,
            "Bean Factory": 12,
            "Public Method": 13,
            "Unknown": 14,
        }
        return priority_map.get(entry_type, 14)

    def search_methods(self, keyword: str):
        """キーワードでメソッドを検索"""
        print(f"\n検索結果: '{keyword}'")
        print(f"{'=' * 80}\n")

        matches = []
        for method in self.method_info.keys():
            if keyword.lower() in method.lower():
                info = self.method_info[method]
                matches.append((method, info.get("class", "")))

        if not matches:
            print("該当するメソッドが見つかりませんでした", file=sys.stderr)
            return

        for i, (method, class_name) in enumerate(matches, 1):
            print(f"{i}. {method}")
            print(f"   クラス: {class_name}")
            print()

    def extract_sql_to_files(self, output_dir: str = "./found_sql") -> None:
        """
        SQL文を抽出してファイルに出力

        Args:
            output_dir: SQL出力先ディレクトリ
        """
        import os
        from pathlib import Path

        # 出力ディレクトリを作成
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # メソッド毎にSQL文を集約
        method_sqls: Dict[str, List[str]] = defaultdict(list)

        for method, info in self.method_info.items():
            sql = info.get("sql", "")
            if sql and sql.strip():
                # SQL文を分割 (複数ある場合は " ||| " で連結されている)
                sqls = [s.strip() for s in sql.split(" ||| ") if s.strip()]
                method_sqls[method].extend(sqls)

        if not method_sqls:
            print("SQL文が見つかりませんでした", file=sys.stderr)
            return

        # ファイル出力
        file_count = 0
        for method, sqls in method_sqls.items():
            safe_name = self._sanitize_method_name(method)

            if len(sqls) == 1:
                filename = f"{safe_name}.sql"
                self._write_sql_file(os.path.join(output_dir, filename), sqls[0])
                file_count += 1
            else:
                for idx, sql in enumerate(sqls, 1):
                    filename = f"{safe_name}_{idx}.sql"
                    self._write_sql_file(os.path.join(output_dir, filename), sql)
                    file_count += 1

        print(f"\n{file_count} 個のSQLファイルを {output_dir} に出力しました")

    def _sanitize_method_name(self, method_signature: str) -> str:
        """
        メソッドシグネチャをファイル名として安全な文字列に変換

        Args:
            method_signature: メソッドシグネチャ (例: "com.example.Class#method()")

        Returns:
            ファイル名として安全な文字列
        """
        name = method_signature.replace(
            "#", "."
        )  # メソッドとクラスの区切りをドットに変換
        # Windowsでファイル名として安全でない文字をアンダースコアに変換
        name = re.sub(r'[<>:"/\\|?*]', "_", name)
        name = re.sub(r"_+", "_", name)  # 連続するアンダースコアを1つに
        name = name.strip("_")
        return name[:200]  # ファイル名の長さを制限

    def _split_by_comma_outside_parens(self, text: str) -> list[str]:
        """
        括弧の外側にあるカンマのみで文字列を分割する。
        関数内のカンマ(例: COALESCE(a, b))では分割しない。

        Args:
            text: 分割対象の文字列

        Returns:
            分割された文字列のリスト
        """
        parts = []
        current_part = []
        paren_depth = 0

        i = 0
        while i < len(text):
            char = text[i]

            if char == "(":
                paren_depth += 1
                current_part.append(char)
            elif char == ")":
                paren_depth -= 1
                current_part.append(char)
            elif char == "," and paren_depth == 0:
                # 括弧の外側のカンマを発見
                # カンマの後のスペースもスキップ
                parts.append("".join(current_part))
                current_part = []
                # カンマの後のスペースをスキップ
                i += 1
                while i < len(text) and text[i] == " ":
                    i += 1
                continue
            else:
                current_part.append(char)

            i += 1

        # 最後の部分を追加
        if current_part:
            parts.append("".join(current_part))

        return parts

    def _format_sql(self, sql_text: str) -> str:
        """
        SQL文を整形（カスタムルール適用）
        - SELECTやFROMなどのキーワードの後は改行してインデント
        - サブクエリーなどの部分はインデントをネスト
        - カラムやテーブルの指定は1行に1つ

        Args:
            sql_text: 整形前のSQL文

        Returns:
            整形後のSQL文
        """
        try:
            import sqlparse

            # 0. コメント除去 (/* ... */)
            # sqlparseの整形前に除去しないと、整形によってコメントの位置がおかしくなる可能性があるため
            sql_text = re.sub(r"/\*[\s\S]*?\*/", "", sql_text)

            # 1. sqlparseで基本整形
            formatted = sqlparse.format(
                sql_text,
                reindent=True,
                keyword_case="upper",
                indent_width=2,
                wrap_after=80,
            )

            # 2. カスタムルール適用（キーワード後の改行とインデント調整）
            lines = formatted.splitlines()
            new_lines = []
            keywords = [
                "SELECT",
                "FROM",
                "WHERE",
                "GROUP BY",
                "ORDER BY",
                "HAVING",
                "SET",
                "VALUES",
                "JOIN",
                "LEFT JOIN",
                "RIGHT JOIN",
                "INNER JOIN",
                "OUTER JOIN",
            ]

            current_alignment_indent = None
            replacement_indent = None

            for line in lines:
                # 整形によるインデント調整ブロック内かチェック
                if current_alignment_indent is not None:
                    if line.startswith(current_alignment_indent):
                        content = line[len(current_alignment_indent) :]

                        # 括弧の外側のカンマのみで分割して1行1つにする
                        parts = self._split_by_comma_outside_parens(content)
                        for i, part in enumerate(parts):
                            if i < len(parts) - 1:
                                new_lines.append(replacement_indent + part + ",")
                            else:
                                new_lines.append(replacement_indent + part)
                        continue
                    else:
                        # インデントが変わったのでブロック終了
                        current_alignment_indent = None
                        replacement_indent = None

                stripped = line.lstrip()
                base_indent = line[: len(line) - len(stripped)]

                matched_keyword = None
                prefix = ""

                # キーワード判定（(SELECT のようなケースも考慮）
                for kw in keywords:
                    if stripped.startswith(kw + " "):
                        matched_keyword = kw
                        prefix = ""
                        break
                    if stripped.startswith("(" + kw + " "):
                        matched_keyword = kw
                        prefix = "("
                        break

                if matched_keyword:
                    # コンテンツは "PREFIX KEYWORD " の後ろから
                    start_index = len(prefix) + len(matched_keyword) + 1
                    content = stripped[start_index:]

                    if content.strip():
                        # キーワード行を出力
                        new_lines.append(base_indent + prefix + matched_keyword)

                        # 括弧の外側のカンマのみで分割して1行1つにする
                        parts = self._split_by_comma_outside_parens(content)

                        # 新しいインデントはベース + 2スペース
                        new_indent = base_indent + "  "

                        for i, part in enumerate(parts):
                            if i < len(parts) - 1:
                                new_lines.append(new_indent + part + ",")
                            else:
                                new_lines.append(new_indent + part)

                        # sqlparseの整形で揃えられた後続行をキャッチするためのインデント長
                        # sqlparseはコンテンツの開始位置に合わせてインデントする
                        align_len = len(base_indent) + start_index
                        current_alignment_indent = " " * align_len
                        replacement_indent = new_indent
                    else:
                        new_lines.append(line)
                else:
                    new_lines.append(line)

            return "\n".join(new_lines)

        except ImportError:
            print(
                "警告: sqlparseがインストールされていません。SQL整形をスキップします。",
                file=sys.stderr,
            )
            print("  インストール: pip install sqlparse", file=sys.stderr)
            return sql_text
        except Exception as e:
            print(f"警告: SQL整形中にエラーが発生しました: {e}", file=sys.stderr)
            return sql_text

    def _write_sql_file(self, filepath: str, sql_text: str) -> None:
        """
        SQL文をファイルに書き込み

        Args:
            filepath: 出力ファイルパス
            sql_text: SQL文
        """
        try:
            formatted_sql = self._format_sql(sql_text)
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(formatted_sql)
        except Exception as e:
            print(
                f"警告: SQLファイルの書き込みに失敗しました ({filepath}): {e}",
                file=sys.stderr,
            )

    def analyze_table_usage(
        self, sql_dir: str = "./found_sql", table_list_file: str = "./table_list.tsv"
    ) -> None:
        """
        SQLファイルから使用テーブルを検出

        Args:
            sql_dir: SQLファイルが格納されているディレクトリ
            table_list_file: テーブル一覧TSVファイルのパス
        """
        import os
        from pathlib import Path

        # テーブル一覧を読み込み
        if not os.path.exists(table_list_file):
            print(
                f"エラー: テーブル一覧ファイルが見つかりません: {table_list_file}",
                file=sys.stderr,
            )
            return

        table_list = self._load_table_list(table_list_file)

        if not table_list:
            print(f"警告: テーブル一覧が空です", file=sys.stderr)
            return

        # SQLファイルを走査
        sql_dir_path = Path(sql_dir)
        if not sql_dir_path.exists():
            print(
                f"エラー: SQLディレクトリが見つかりません: {sql_dir}", file=sys.stderr
            )
            return

        sql_files = sorted(sql_dir_path.glob("*.sql"))

        if not sql_files:
            print(f"警告: SQLファイルが見つかりません: {sql_dir}", file=sys.stderr)
            return

        #  clipでコピーした結果をExcelに貼り付けられるにはShift_JISで出力する
        sys.stdout.reconfigure(encoding=self.output_tsv_encoding)

        # ヘッダーを出力
        print("SQLファイル名\t物理テーブル名\t論理テーブル名\t補足情報")

        # 各SQLファイルを解析
        for sql_file in sql_files:
            try:
                with open(sql_file, "r", encoding="utf-8") as f:
                    sql_content = f.read()

                # テーブルを検出
                found_tables = self._find_tables_in_sql(sql_content, table_list)

                # 結果を出力
                if found_tables:
                    for physical_name, logical_name, note in found_tables:
                        print(
                            f"{sql_file.name}\t{physical_name}\t{logical_name}\t{note}"
                        )
                else:
                    # テーブルが見つからない場合も1行出力
                    print(f"{sql_file.name}\t\t\t")

            except Exception as e:
                print(
                    f"エラー: SQLファイルの読み込みに失敗しました ({sql_file.name}): {e}",
                    file=sys.stderr,
                )

    def _load_table_list(self, table_list_file: str) -> List[tuple[str, str, str]]:
        """
        テーブル一覧をTSVファイルから読み込み

        Args:
            table_list_file: テーブル一覧TSVファイルのパス

        Returns:
            (物理テーブル名, 論理テーブル名, 補足情報)のリスト
        """
        table_list = []

        try:
            with open(table_list_file, "r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter="\t")
                for row in reader:
                    # コメント行や空行をスキップ
                    if not row or (row[0] and row[0].startswith("#")):
                        continue

                    # 最低でも物理テーブル名が必要
                    if len(row) >= 1 and row[0].strip():
                        physical_name = row[0].strip()
                        logical_name = row[1].strip() if len(row) >= 2 else ""
                        note = row[2].strip() if len(row) >= 3 else ""
                        table_list.append((physical_name, logical_name, note))

        except Exception as e:
            print(
                f"エラー: テーブル一覧ファイルの読み込みに失敗しました: {e}",
                file=sys.stderr,
            )

        return table_list

    def _find_tables_in_sql(
        self, sql_content: str, table_list: List[tuple[str, str, str]]
    ) -> List[tuple[str, str, str]]:
        """
        SQL文からテーブルを検出

        Args:
            sql_content: SQL文
            table_list: テーブル一覧

        Returns:
            検出されたテーブル情報のリスト
        """
        # SQL文を大文字化して検索
        sql_upper = sql_content.upper()

        found_tables = []
        seen_tables = set()  # 重複を避けるため

        for physical_name, logical_name, note in table_list:
            # テーブル名を大文字化して検索
            table_upper = physical_name.upper()

            # 単語境界を考慮してテーブル名を検索
            # テーブル名の前後が英数字でないことを確認
            pattern = r"\b" + re.escape(table_upper) + r"\b"

            if re.search(pattern, sql_upper):
                if physical_name not in seen_tables:
                    found_tables.append((physical_name, logical_name, note))
                    seen_tables.add(physical_name)

        return found_tables

    def _extract_method_signature_parts(self, method_signature: str) -> Dict[str, str]:
        """
        メソッドシグネチャを分解

        Args:
            method_signature: メソッドシグネチャ (例: "com.example.service.UserService#getUser(String)")

        Returns:
            各要素を含む辞書
        """
        if "#" not in method_signature:
            return {
                "package": "",
                "class": "",
                "simple_class": "",
                "method": method_signature,
                "full_signature": method_signature,
            }

        class_part, method_part = method_signature.split("#", 1)

        # パッケージ名とクラス名を分離
        if "." in class_part:
            package = ".".join(class_part.split(".")[:-1])
            simple_class = class_part.split(".")[-1]
        else:
            package = ""
            simple_class = class_part

        return {
            "package": package,
            "class": class_part,
            "simple_class": simple_class,
            "method": method_part,
            "full_signature": method_signature,
        }

    def _format_tree_display(self, method_signature: str) -> str:
        """
        L列以降に表示するツリー用のメソッド名を生成
        パッケージ名を除外し、SimpleClassName#methodName(params)形式
        引数のクラス名からもパッケージ名を除去

        Args:
            method_signature: メソッドシグネチャ

        Returns:
            ツリー表示用のメソッド名
        """
        # _shorten_method_signatureを再利用
        return self._shorten_method_signature(method_signature)

    def _collect_tree_data(
        self,
        root_method: str,
        max_depth: int,
        follow_implementations: bool,
        visited: Optional[Set[str]] = None,
        depth: int = 0,
        parent_relation: str = "",
        accumulated_instances: Optional[Set[str]] = None,  # 累積されたインスタンス情報
        max_depth_reached: Optional[List[bool]] = None,  # 最大深度到達フラグ
    ) -> List[Dict[str, any]]:
        """
        1つの呼び出しツリーを再帰的にトラバースし、全メソッド情報を収集

        Args:
            root_method: ルートメソッド
            max_depth: 最大深度
            follow_implementations: 実装クラス候補を追跡するか
            visited: 訪問済みメソッド集合（循環参照チェック用）
            depth: 現在の深度
            parent_relation: 呼び出し種別（"親クラスメソッド" / "インターフェース" / "実装クラス候補" / ""）
            accumulated_instances: 呼び出しツリーの上位から累積された生成インスタンス情報

        Returns:
            各メソッドの情報を含む辞書のリスト
        """
        if visited is None:
            visited = set()

        result = []

        if depth > max_depth:
            # 最大深度に到達した場合、フラグをセット
            if max_depth_reached is not None:
                max_depth_reached[0] = True
            return result

        # 除外ルールチェック
        if not self.exclusion_manager.should_include(root_method):
            return result

        # 現在のメソッドで生成されるインスタンスを収集し、累積に追加
        current_instances = self._collect_created_instances(root_method)
        if accumulated_instances is None:
            accumulated_instances = current_instances
        else:
            accumulated_instances.update(current_instances)

        # 循環参照チェック
        is_circular = root_method in visited

        # メソッド情報を取得
        info = self.method_info.get(root_method, {})
        parts = self._extract_method_signature_parts(root_method)

        # 現在のメソッドを結果に追加
        result.append(
            {
                "depth": depth,
                "method": root_method,
                "package": parts["package"],
                "class": parts["class"],
                "simple_method": parts["method"],
                "javadoc": info.get("javadoc", ""),
                "parent_relation": parent_relation,
                "sql": info.get("sql", ""),
                "is_circular": is_circular,
                "tree_display": self._format_tree_display(root_method),
                "httpCalls": info.get("httpCalls", []),  # HTTPクライアント呼び出し情報
                "hit_words": info.get("hit_words", ""),  # 検出ワード
            }
        )

        # 循環参照の場合は子ノードを展開しない
        if is_circular:
            return result

        visited_copy = visited.copy()
        visited_copy.add(root_method)

        # 除外ルールで配下を除外する場合
        if self.exclusion_manager.should_exclude_children(root_method):
            return result

        # 子ノードを再帰的に処理
        callees = self.forward_calls.get(root_method, [])
        for callee_info in callees:
            callee = callee_info["method"]

            # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
            if not self.exclusion_manager.should_include(callee):
                continue

            # 呼び出し種別を判定
            # 1. 親クラスのメソッドの場合: 親クラス
            # 2. インターフェースのメソッドの場合: 実装クラス側で「インターフェース」を設定
            if callee_info["is_parent_method"] == "Yes":
                relation = "親クラスメソッド"
            elif callee_info["implementations"]:
                # 実装がある＝インターフェースまたは抽象クラスのメソッド
                relation = "インターフェース"
            else:
                relation = ""

            # 呼び出し先を再帰的に収集
            result.extend(
                self._collect_tree_data(
                    callee,
                    max_depth,
                    follow_implementations,
                    visited_copy.copy(),
                    depth + 1,
                    relation,
                    accumulated_instances,  # 累積インスタンスを渡す
                    max_depth_reached,  # 最大深度到達フラグを渡す
                )
            )

            # 実装クラス候補がある場合
            if follow_implementations and callee_info["implementations"]:
                implementations = [
                    impl.strip().split(" ")[0]
                    for impl in callee_info["implementations"].split(",")
                    if impl.strip()
                ]

                # 累積されたインスタンス情報に基づいてフィルタリング
                if accumulated_instances:
                    filtered_implementations = (
                        self._filter_implementations_by_accumulated_instances(
                            accumulated_instances, implementations
                        )
                    )
                else:
                    filtered_implementations = implementations

                for impl_class in filtered_implementations:
                    impl_method = self._find_implementation_method(callee, impl_class)
                    if impl_method:
                        # Iモード: 除外対象の場合、ノード自体を表示せずスキップ
                        if not self.exclusion_manager.should_include(impl_method):
                            continue

                        result.extend(
                            self._collect_tree_data(
                                impl_method,
                                max_depth,
                                follow_implementations,
                                visited_copy.copy(),
                                depth + 1,
                                "実装クラス候補",
                                accumulated_instances,  # 累積インスタンスを渡す
                                max_depth_reached,  # 最大深度到達フラグを渡す
                            )
                        )

        return result

    def export_tree_to_csv(
        self,
        entry_points_file: Optional[str],
        output_file: Optional[str],
        max_depth: int = 20,
        follow_implementations: bool = True,
    ) -> None:
        """
        CSV形式でエントリーポイントからの呼び出しメソッド一覧をエクスポート

        Args:
            entry_points_file: エントリーポイントファイル（Noneの場合は厳密モードのエントリーポイント）
            output_file: 出力ファイル名（Noneの場合は標準出力）
            max_depth: 最大深度
            follow_implementations: 実装クラス候補を追跡するか
        """
        # エントリーポイントの決定
        entry_points: List[str] = []

        if entry_points_file:
            try:
                with open(entry_points_file, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        # 空行とコメント行をスキップ
                        if line and not line.startswith("#"):
                            entry_points.append(line)
            except Exception as e:
                print(
                    f"エラー: エントリーポイントファイルの読み込みに失敗しました: {e}",
                    file=sys.stderr,
                )
                return
        else:
            # 厳密モードのエントリーポイントを取得
            all_callees = set()
            for callees in self.forward_calls.values():
                for callee_info in callees:
                    all_callees.add(callee_info["method"])

            for method, info in self.method_info.items():
                if method not in all_callees and info.get("is_entry_point"):
                    entry_points.append(method)

        if not entry_points:
            print("警告: エントリーポイントが見つかりませんでした", file=sys.stderr)
            return

        # CSVヘッダー
        headers = [
            "エントリーポイント（fully qualified name）",
            "エントリーポイントのパッケージ名",
            "エントリーポイントのクラス名",
            "エントリーポイントのメソッド名",
            "呼び出しメソッド（fully qualified name）",
            "呼び出しメソッドのパッケージ名",
            "呼び出しメソッドのクラス名",
            "呼び出しメソッドのメソッド名",
        ]

        def extract_method_name_only(method_with_params: str) -> str:
            """メソッド名から引数を除去する"""
            # method(params) -> method
            paren_pos = method_with_params.find("(")
            if paren_pos >= 0:
                return method_with_params[:paren_pos]
            return method_with_params

        # 出力先を決定
        if output_file:
            try:
                f = open(output_file, "w", encoding="Shift_JIS", newline="")
            except Exception as e:
                print(f"エラー: ファイルを開けません: {e}", file=sys.stderr)
                return
        else:
            # 標準出力の場合もShift_JISにreconfigure
            sys.stdout.reconfigure(encoding="Shift_JIS")
            f = sys.stdout

        try:
            writer = csv.writer(f)
            writer.writerow(headers)

            # 最大深度に到達したエントリーポイントを追跡
            max_depth_reached_entries: List[str] = []

            for entry_point in entry_points:
                # エントリーポイント自身の情報を分解
                ep_parts = self._extract_method_signature_parts(entry_point)
                ep_method_name = extract_method_name_only(ep_parts["method"])

                # 最大深度到達フラグを初期化
                max_depth_reached_flag: List[bool] = [False]

                # ツリーデータを収集
                tree_data = self._collect_tree_data(
                    entry_point,
                    max_depth,
                    follow_implementations,
                    max_depth_reached=max_depth_reached_flag,
                )

                # 最大深度に到達した場合、エントリーポイントを記録
                if max_depth_reached_flag[0]:
                    max_depth_reached_entries.append(entry_point)

                # 各メソッドについてCSV行を出力
                for node in tree_data:
                    callee_method_name = extract_method_name_only(node["simple_method"])

                    row = [
                        entry_point,  # エントリーポイント（fully qualified name）
                        ep_parts["package"],  # エントリーポイントのパッケージ名
                        ep_parts["simple_class"],  # エントリーポイントのクラス名
                        ep_method_name,  # エントリーポイントのメソッド名
                        node["method"],  # 呼び出しメソッド（fully qualified name）
                        node["package"],  # 呼び出しメソッドのパッケージ名
                        (
                            node["class"].split(".")[-1] if node["class"] else ""
                        ),  # 呼び出しメソッドのクラス名
                        callee_method_name,  # 呼び出しメソッドのメソッド名
                    ]
                    writer.writerow(row)

            if output_file:
                print(f"CSVを {output_file} にエクスポートしました", file=sys.stderr)

            # 最大深度に到達したエントリーポイントの警告を出力
            if max_depth_reached_entries:
                print(
                    f"\n警告: 以下のエントリーポイントは最大深度({max_depth})に到達しました。"
                    "ツリーが切り捨てられている可能性があります:",
                    file=sys.stderr,
                )
                for ep in max_depth_reached_entries:
                    print(f"  - {ep}", file=sys.stderr)
                print(
                    f"ヒント: --depth オプションで深度を増やすことを検討してください。",
                    file=sys.stderr,
                )
        finally:
            if output_file and f != sys.stdout:
                f.close()

    def export_tree_to_excel(
        self,
        entry_points_file: Optional[str],
        output_file: str,
        max_depth: int = 20,
        follow_implementations: bool = True,
        include_tree: bool = True,
        include_sql: bool = True,
    ) -> None:
        """
        Excel形式でツリーをエクスポート

        Args:
            entry_points_file: エントリーポイントファイル（Noneの場合は厳密モードのエントリーポイント）
            output_file: 出力ファイル名
            max_depth: 最大深度
            follow_implementations: 実装クラス候補を追跡するか
            include_tree: L列以降の呼び出しツリーを出力するか
            include_sql: AZ列のSQL文を出力するか
        """
        # エントリーポイントの決定
        entry_points: List[str] = []

        if entry_points_file:
            try:
                with open(entry_points_file, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        # 空行とコメント行をスキップ
                        if line and not line.startswith("#"):
                            entry_points.append(line)
            except Exception as e:
                print(
                    f"エラー: エントリーポイントファイルの読み込みに失敗しました: {e}",
                    file=sys.stderr,
                )
                return
        else:
            # 厳密モードのエントリーポイントを取得
            all_callees = set()
            for callees in self.forward_calls.values():
                for callee_info in callees:
                    all_callees.add(callee_info["method"])

            for method, info in self.method_info.items():
                if method not in all_callees and info.get("is_entry_point"):
                    entry_points.append(method)

        if not entry_points:
            print("警告: エントリーポイントが見つかりませんでした", file=sys.stderr)
            return

        print(f"エントリーポイント数: {len(entry_points)}")

        # Excelワークブックの作成
        wb = openpyxl.Workbook()
        ws = wb.active
        if not isinstance(ws, openpyxl.worksheet.worksheet.Worksheet):
            raise TypeError("Active sheet is not a Worksheet")

        # 背景色（薄めのオリーブ）と罫線（破線）を定義
        olive_fill = PatternFill(
            start_color="C4D79B", end_color="C4D79B", fill_type="solid"
        )
        dashed_border = Border(
            left=Side(style="dashed", color="000000"),
            right=Side(style="dashed", color="000000"),
            top=Side(style="dashed", color="000000"),
            bottom=Side(style="dashed", color="000000"),
        )

        # NamedStyleを作成（フォントとアライメントを定義）
        default_style = NamedStyle(name="default_style")
        default_style.font = Font(name="Meiryo UI")
        default_style.alignment = Alignment(vertical="center", horizontal="left")
        default_style.border = dashed_border

        green_style = NamedStyle(name="green_style")
        green_style.font = Font(name="Meiryo UI", color="008000")
        green_style.alignment = Alignment(vertical="center", horizontal="left")
        green_style.border = dashed_border

        # ヘッダ用スタイル（オリーブ背景色）
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name="Meiryo UI")
        header_style.alignment = Alignment(vertical="center", horizontal="left")
        header_style.fill = olive_fill
        header_style.border = dashed_border

        # L列用スタイル（太字）
        tree_style = NamedStyle(name="tree_style")
        tree_style.font = Font(name="Meiryo UI", bold=True)
        tree_style.alignment = Alignment(vertical="center", horizontal="left")
        tree_style.border = dashed_border

        # インターフェース用スタイル（斜体、グレー）
        interface_style = NamedStyle(name="interface_style")
        interface_style.font = Font(name="Meiryo UI", italic=True, color="808080")
        interface_style.alignment = Alignment(vertical="center", horizontal="left")
        interface_style.border = dashed_border

        # 実装クラス候補用スタイル（下線）
        impl_style = NamedStyle(name="impl_style")
        impl_style.font = Font(name="Meiryo UI", underline="single")
        impl_style.alignment = Alignment(vertical="center", horizontal="left")
        impl_style.border = dashed_border

        # F列（呼び出し種別）用スタイル（縮小して全体を表示）
        shrink_style = NamedStyle(name="shrink_style")
        shrink_style.font = Font(name="Meiryo UI")
        shrink_style.alignment = Alignment(
            vertical="center", horizontal="left", shrink_to_fit=True
        )
        shrink_style.border = dashed_border

        # スタイルをワークブックに登録
        wb.add_named_style(default_style)
        wb.add_named_style(green_style)
        wb.add_named_style(header_style)
        wb.add_named_style(tree_style)
        wb.add_named_style(interface_style)
        wb.add_named_style(impl_style)
        wb.add_named_style(shrink_style)

        # C～E、G列の幅を30に設定
        for col_letter in ["C", "D", "E", "G"]:
            ws.column_dimensions[col_letter].width = 30

        # L列以降の列幅を5に設定
        tree_start_col = column_index_from_string("L")
        tree_end_col = tree_start_col + max_depth - 1  # 呼び出しツリーの最終列
        for col_idx in range(tree_start_col, tree_start_col + max_depth):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 5

        # --depthオプションに基づく動的列計算
        # L列からmax_depth列目がSQL有無列になる（新しい順序）
        sql_exists_col = tree_start_col + max_depth  # AF列 (depth=20の場合)
        sql_content_col = sql_exists_col + 1  # AG列 (depth=20の場合)
        http_exists_col = sql_content_col + 1  # AH列 (depth=20の場合)
        http_request_col = http_exists_col + 1  # AI列 (depth=20の場合)
        hitwords_col = http_request_col + 1  # AJ列 (depth=20の場合)

        # 1行目: L1に「呼び出しツリー」を出力
        if include_tree:
            ws.cell(row=1, column=tree_start_col, value="呼び出しツリー").style = (
                "header_style"
            )

        # 2行目: ヘッダ行
        header_row = 2
        #   A～G列
        ws.cell(row=header_row, column=1, value="エントリーポイント").style = (
            "header_style"
        )
        ws.cell(row=header_row, column=2, value="呼び出しメソッド").style = (
            "header_style"
        )
        ws.cell(row=header_row, column=3, value="パッケージ名").style = "header_style"
        ws.cell(row=header_row, column=4, value="クラス名").style = "header_style"
        ws.cell(row=header_row, column=5, value="メソッド名").style = "header_style"
        ws.cell(row=header_row, column=6, value="呼び出し種別").style = "header_style"
        ws.cell(row=header_row, column=7, value="Javadoc").style = "header_style"
        #   L2～呼び出しツリー最終列に連番（1,2,3...）
        if include_tree:
            for i, col_idx in enumerate(
                range(tree_start_col, tree_end_col + 1), start=1
            ):
                ws.cell(row=header_row, column=col_idx, value=i).style = "header_style"
        #   動的列: SQL有無、SQL文（include_sqlがTrueの場合のみ）
        if include_sql:
            ws.cell(row=header_row, column=sql_exists_col, value="SQL有無").style = (
                "header_style"
            )
            ws.cell(row=header_row, column=sql_content_col, value="SQL文").style = (
                "header_style"
            )
        #   動的列: HTTP有無、HTTPリクエスト
        ws.cell(row=header_row, column=http_exists_col, value="HTTP有無").style = (
            "header_style"
        )
        ws.cell(
            row=header_row, column=http_request_col, value="HTTPリクエスト"
        ).style = "header_style"
        #   動的列: hitWords列
        ws.cell(row=header_row, column=hitwords_col, value="検出ワード").style = (
            "header_style"
        )

        current_row = 3  # データは3行目から

        # 最大深度に到達したエントリーポイントを追跡
        max_depth_reached_entries: List[str] = []

        for entry_point in entry_points:
            print(f"処理中: {entry_point}")

            # 最大深度到達フラグを初期化
            max_depth_reached_flag: List[bool] = [False]

            # ツリーデータを収集
            tree_data = self._collect_tree_data(
                entry_point,
                max_depth,
                follow_implementations,
                max_depth_reached=max_depth_reached_flag,
            )

            # 最大深度に到達した場合、エントリーポイントを記録
            if max_depth_reached_flag[0]:
                max_depth_reached_entries.append(entry_point)

            # Excelに書き込み
            for node in tree_data:
                # A列: エントリーポイント（すべての行に出力）
                ws.cell(row=current_row, column=1, value=entry_point).style = (
                    "default_style"
                )

                # B列: 呼び出しメソッド（fully qualified name）
                ws.cell(row=current_row, column=2, value=node["method"]).style = (
                    "default_style"
                )

                # C列: パッケージ名
                ws.cell(row=current_row, column=3, value=node["package"]).style = (
                    "default_style"
                )

                # D列: クラス名（パッケージ名を除いたシンプルなクラス名）
                simple_class = node["class"].split(".")[-1] if node["class"] else ""
                ws.cell(row=current_row, column=4, value=simple_class).style = (
                    "default_style"
                )

                # E列: メソッド名（simple name）
                ws.cell(
                    row=current_row, column=5, value=node["simple_method"]
                ).style = "default_style"

                # F列: 呼び出し種別（親クラス / インターフェース / 実装クラス）、空の場合は半角スペース
                parent_relation_value = (
                    node["parent_relation"] if node["parent_relation"] else " "
                )
                ws.cell(
                    row=current_row, column=6, value=parent_relation_value
                ).style = "shrink_style"

                # G列: Javadoc（緑フォント）、空の場合は半角スペース
                javadoc_value = node["javadoc"] if node["javadoc"] else " "
                ws.cell(row=current_row, column=7, value=javadoc_value).style = (
                    "green_style"
                )

                # L列以降: 呼び出しツリー（include_treeがTrueの場合のみ）
                if include_tree:
                    tree_col = tree_start_col + node["depth"]
                    tree_text = str(node["tree_display"] or "")
                    if node["is_circular"] and tree_text:
                        tree_text = tree_text + " [循環参照]"
                    # L列（depth=0）は太字、インターフェースは斜体、実装クラス候補は下線
                    if tree_col == tree_start_col:
                        ws.cell(
                            row=current_row, column=tree_col, value=tree_text
                        ).style = "tree_style"
                    elif node["parent_relation"] == "インターフェース":
                        ws.cell(
                            row=current_row, column=tree_col, value=tree_text
                        ).style = "interface_style"
                    elif node["parent_relation"] == "実装クラス候補":
                        ws.cell(
                            row=current_row, column=tree_col, value=tree_text
                        ).style = "impl_style"
                    else:
                        ws.cell(
                            row=current_row, column=tree_col, value=tree_text
                        ).style = "default_style"

                # 動的列: SQL有無、SQL文（include_sqlがTrueの場合のみ）
                if include_sql:
                    sql_marker = "●" if node["sql"] else ""
                    ws.cell(
                        row=current_row, column=sql_exists_col, value=sql_marker
                    ).style = "default_style"
                    if node["sql"]:
                        ws.cell(
                            row=current_row, column=sql_content_col, value=node["sql"]
                        ).style = "default_style"

                # 動的列: HTTP有無、HTTPリクエスト
                http_calls = node.get("httpCalls", [])
                http_marker = "●" if http_calls else ""
                ws.cell(
                    row=current_row, column=http_exists_col, value=http_marker
                ).style = "default_style"

                if http_calls:
                    http_details = ", ".join(
                        f"{call.get('httpMethod', 'UNKNOWN')} - {call.get('uri', '${UNRESOLVED}')}"
                        for call in http_calls
                    )
                    ws.cell(
                        row=current_row, column=http_request_col, value=http_details
                    ).style = "default_style"

                # 動的列: hitWords
                hit_words = node.get("hit_words", "")
                if hit_words:
                    ws.cell(
                        row=current_row, column=hitwords_col, value=hit_words
                    ).style = "default_style"

                current_row += 1

        # フィルターを設定（A2:AO<最終行>、ヘッダは2行目）
        last_row = current_row - 1
        ao_col = column_index_from_string("AO")
        filter_range = f"A2:{get_column_letter(ao_col)}{last_row}"
        ws.auto_filter.ref = filter_range

        # 1行目の全セルにヘッダースタイルを適用
        for col_idx in range(1, ao_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.style != "header_style":
                cell.style = "header_style"

        # 2行目～最終行の全セルに書式を適用（未設定のセルのみ）
        for row_idx in range(2, last_row + 1):
            for col_idx in range(1, ao_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # スタイルが未設定（Normal）の場合のみ適用
                if cell.style == "Normal" or cell.style is None:
                    if row_idx == 2:
                        cell.style = "header_style"
                    else:
                        cell.style = "default_style"

        # 条件付き書式: L列に値がある場合は行全体の背景色をライトグレーに
        light_gray_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        # データ範囲（3行目～最終行）に条件付き書式を適用
        data_range = f"A3:{get_column_letter(ao_col)}{last_row}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'$L3<>""'],
                fill=light_gray_fill,
            ),
        )

        # ウィンドウ枚の固定（A3セルで固定）
        ws.freeze_panes = "A3"

        # Excelファイルの保存
        try:
            wb.save(output_file)
            print(f"ツリーを {output_file} にエクスポートしました")
            print(f"総行数: {current_row - 1}")
        except Exception as e:
            print(f"エラー: Excelファイルの保存に失敗しました: {e}", file=sys.stderr)

        # 最大深度に到達したエントリーポイントの警告を出力
        if max_depth_reached_entries:
            print(
                f"\n警告: 以下のエントリーポイントは最大深度({max_depth})に到達しました。"
                "ツリーが切り捨てられている可能性があります:",
                file=sys.stderr,
            )
            for ep in max_depth_reached_entries:
                print(f"  - {ep}", file=sys.stderr)
            print(
                f"ヒント: --depth オプションで深度を増やすことを検討してください。",
                file=sys.stderr,
            )


# サブコマンドハンドラー関数


def handle_entries(args, visualizer: CallTreeVisualizer) -> None:
    """entriesサブコマンドの処理"""
    if args.tsv:
        visualizer.list_entry_points_tsv(args.min_calls, args.strict)
    else:
        visualizer.list_entry_points(args.min_calls, args.strict)


def handle_search(args, visualizer: CallTreeVisualizer) -> None:
    """searchサブコマンドの処理"""
    visualizer.search_methods(args.keyword)


def handle_forward(args, visualizer: CallTreeVisualizer) -> None:
    """forwardサブコマンドの処理"""
    visualizer.print_forward_tree(
        args.method,
        args.depth,
        show_class=args.show_class,
        show_sql=args.show_sql,
        follow_implementations=args.follow_impl,
        verbose=args.verbose,
        use_tab=args.tab,
        short_mode=args.short,
    )


def handle_reverse(args, visualizer: CallTreeVisualizer) -> None:
    """reverseサブコマンドの処理"""
    visualizer.print_reverse_tree(
        args.method,
        args.depth,
        show_class=args.show_class,
        follow_overrides=args.follow_override,
        verbose=args.verbose,
        use_tab=args.tab,
        short_mode=args.short,
    )


def handle_export(args, visualizer: CallTreeVisualizer) -> None:
    """exportサブコマンドの処理"""
    visualizer.export_tree_to_file(
        args.method,
        args.output_file,
        args.depth,
        args.format,
        args.follow_impl,
    )


def handle_export_excel(args, visualizer: CallTreeVisualizer) -> None:
    """export-excelサブコマンドの処理"""
    visualizer.export_tree_to_excel(
        args.entry_points,
        args.output_file,
        args.depth,
        args.follow_impl,
        args.include_tree,
        args.include_sql,
    )


def handle_export_csv(args, visualizer: CallTreeVisualizer) -> None:
    """export-csvサブコマンドの処理"""
    visualizer.export_tree_to_csv(
        args.entry_points,
        args.output_file,
        args.depth,
        args.follow_impl,
    )


def handle_extract_sql(args, visualizer: CallTreeVisualizer) -> None:
    """extract-sqlサブコマンドの処理"""
    visualizer.extract_sql_to_files(args.output_dir)


def handle_analyze_tables(args, visualizer: CallTreeVisualizer) -> None:
    """analyze-tablesサブコマンドの処理"""
    visualizer.analyze_table_usage(args.sql_dir, args.table_list)


class StructureVisualizer:
    """構造情報（クラス階層、インターフェース実装）を可視化するクラス"""

    def __init__(self, json_file: str):
        self.json_file = json_file
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                self.data = json.load(f)
        except Exception as e:
            print(f"JSONファイルの読み込みに失敗しました: {e}", file=sys.stderr)
            sys.exit(1)

    def print_class_tree(
        self,
        root_filter: Optional[str] = None,
        max_depth: int = 50,
        verbose: bool = False,
    ):
        """クラス階層ツリーを表示"""
        classes = self.data.get("classes", [])

        # データの構築
        class_map = {c["className"]: c for c in classes}
        children_map = defaultdict(list)
        roots = set()

        # 全クラス名をセットに（存在チェック用）
        all_class_names = set(c["className"] for c in classes)

        # 親子関係の構築とルートの特定
        for c in classes:
            class_name = c["className"]
            super_class = c.get("superClass")

            if super_class:
                children_map[super_class].append(class_name)
                # 親クラスがデータセットに含まれていない場合、その親クラスもルート候補とする
                if super_class not in all_class_names:
                    roots.add(super_class)
            else:
                # 親クラスがない場合はルート
                roots.add(class_name)

        # フィルタリング適用
        if root_filter:
            target_roots = [r for r in roots if root_filter in r]
            # ルートそのものでなくても、ツリーの途中にあるクラスも指定できるようにする
            if not target_roots and root_filter in all_class_names:
                target_roots = [root_filter]

            if not target_roots:
                # 部分一致で探す
                target_roots = [c for c in all_class_names if root_filter in c]
        else:
            target_roots = sorted(list(roots))

        if not target_roots:
            print(f"該当するクラスが見つかりませんでした: {root_filter}")
            return

        print(f"\n{'=' * 80}")
        print(f"クラス継承ツリー (JSON: {self.json_file})")
        print(f"{'=' * 80}\n")

        for root in target_roots:
            self._print_class_node_recursive(
                root, 0, max_depth, verbose, class_map, children_map, set()
            )
            print()  # ツリー間の改行

    def _print_class_node_recursive(
        self,
        class_name: str,
        depth: int,
        max_depth: int,
        verbose: bool,
        class_map: Dict[str, Dict],
        children_map: Dict[str, List[str]],
        visited: Set[str],
    ):
        if depth > max_depth:
            return

        if class_name in visited:
            indent = "    " * depth
            print(f"{indent}|-- {class_name} [循環参照]")
            return

        visited.add(class_name)

        indent = "    " * depth
        prefix = "|-- " if depth > 0 else ""

        # 表示情報の構築
        display = f"{indent}{prefix}{class_name}"

        # 詳細情報の表示
        info = class_map.get(class_name)
        if verbose and info:
            extras = []
            if info.get("javadoc"):
                extras.append(f"Javadoc: {info.get('javadoc')}")
            else:
                extras.append("Javadoc: (なし)")

            annotations = info.get("annotations", [])
            if annotations:
                anns_str = ", ".join([f"@{a.split('.')[-1]}" for a in annotations])
                extras.append(f"Annotations: [{anns_str}]")
            else:
                extras.append("Annotations: (なし)")

            if extras:
                # クラス名の右側にタブ区切りで表示
                # javadocとアノテーションの間もタブ区切りで表示
                display += f"\t{'\t'.join(extras)}"

        # 外部親クラスの場合の表示
        if not info and depth == 0:
            display += " [外部親クラス]"

        print(display)

        # 子クラスの表示
        children = sorted(children_map.get(class_name, []))
        for child in children:
            self._print_class_node_recursive(
                child,
                depth + 1,
                max_depth,
                verbose,
                class_map,
                children_map,
                visited.copy(),
            )

    def print_interface_impls(
        self, filter_str: Optional[str] = None, verbose: bool = False
    ):
        """インターフェース実装一覧を表示"""
        interfaces = self.data.get("interfaces", [])

        print(f"\n{'=' * 80}")
        print(f"インターフェース実装一覧 (JSON: {self.json_file})")
        print(f"{'=' * 80}\n")

        count = 0
        for iface in interfaces:
            interface_name = iface["interfaceName"]

            # フィルタリング
            if filter_str and filter_str not in interface_name:
                continue

            implementations = iface.get("implementations", [])
            if not implementations:
                continue

            count += 1
            if verbose:
                # verboseモードではインターフェースの詳細情報も表示
                extras = []
                if iface.get("javadoc"):
                    extras.append(f"Javadoc: {iface.get('javadoc')}")
                else:
                    extras.append("Javadoc: (なし)")
                annotations = iface.get("annotations", [])
                if annotations:
                    anns_str = ", ".join([f"@{a.split('.')[-1]}" for a in annotations])
                    extras.append(f"Annotations: [{anns_str}]")
                else:
                    extras.append("Annotations: (なし)")
                # superInterfaces表示
                super_ifaces = iface.get("superInterfaces", [])
                if super_ifaces:
                    extras.append(f"Extends: [{', '.join(super_ifaces)}]")
                # hitWords表示
                hit_words = iface.get("hitWords", [])
                if hit_words:
                    extras.append(f"HitWords: [{', '.join(hit_words)}]")
                print(f"Interface: {interface_name}\t{'\t'.join(extras)}")
            else:
                print(f"Interface: {interface_name}")

            # 実装クラスをタイプ（direct/indirect）とクラス名でソート
            sorted_impls = sorted(
                implementations, key=lambda x: (x["type"] != "direct", x["className"])
            )  # directが先

            for impl in sorted_impls:
                class_name = impl["className"]
                impl_type = impl["type"]
                type_mark = "[Direct]  " if impl_type == "direct" else "[Indirect]"

                display = f"    {type_mark} {class_name}"

                if verbose:
                    extras = []
                    if impl.get("javadoc"):
                        extras.append(f"Javadoc: {impl.get('javadoc')}")
                    else:
                        extras.append("Javadoc: (なし)")

                    annotations = impl.get("annotations", [])
                    if annotations:
                        anns_str = ", ".join(
                            [f"@{a.split('.')[-1]}" for a in annotations]
                        )
                        extras.append(f"Annotations: [{anns_str}]")
                    else:
                        extras.append("Annotations: (なし)")

                    if extras:
                        display += f"\t{'\t'.join(extras)}"

                print(display)
            print()  # インターフェースごとの空行

        if count == 0 and filter_str:
            print(f"該当するインターフェースが見つかりませんでした: {filter_str}")


def handle_class_tree(args) -> None:
    """class-treeサブコマンドの処理"""
    visualizer = StructureVisualizer(args.input_file)
    # root_filterかfilterのどちらかを使用
    f = args.root_filter if args.root_filter else args.filter
    visualizer.print_class_tree(root_filter=f, verbose=args.verbose)


def handle_interface_impls(args) -> None:
    """interface-implsサブコマンドの処理"""
    visualizer = StructureVisualizer(args.input_file)
    visualizer.print_interface_impls(filter_str=args.filter_str, verbose=args.verbose)


def main():
    """メイン関数"""
    import argparse

    # メインパーサーの作成
    parser = argparse.ArgumentParser(
        description="呼び出しツリー可視化スクリプト - JSONファイルから可視化を行います",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
除外ルールファイルのフォーマット:
  <クラス名 or メソッド名><TAB><I|E>
  I: 対象自体を除外
  E: 対象は表示するが、配下の呼び出しを除外

テーブルリストファイル (table_list.tsv) のフォーマット:
  <物理テーブル名><TAB><論理テーブル名><TAB><補足情報>

使用例:
  %(prog)s entries
  %(prog)s entries --no-strict --min-calls 5
  %(prog)s forward 'com.example.Main#main(String[])'
  %(prog)s reverse 'com.example.Service#process()'
  %(prog)s export 'com.example.Main#main(String[])' tree.html --format html
  %(prog)s export-excel call_trees.xlsx --entry-points entry_points.txt
  %(prog)s extract-sql --output-dir ./output/sqls
  %(prog)s analyze-tables --sql-dir ./output/sqls
  %(prog)s class-tree --filter 'com.example'
  %(prog)s interface-impls --interface 'MyService'
        """,
    )

    parser.add_argument(
        "-i",
        "--input",
        dest="input_file",
        default="analyzed_result.json",
        help="入力ファイル（JSONまたはTSV）のパス (デフォルト: analyzed_result.json)",
    )
    parser.add_argument(
        "--exclusion-file",
        help="除外ルールファイルのパス (デフォルト: exclusion_rules.txt)",
    )
    parser.add_argument(
        "--output-tsv-encoding",
        default="Shift_JIS",
        help="出力するTSVのエンコーディング (デフォルト: Shift_JIS (Excelへの貼付けを考慮))",
    )
    parser.add_argument(
        "-d",
        "--debug",
        action="store_true",
        dest="debug_mode",
        help="デバッグモード（インスタンス収集情報を出力）",
    )

    # サブコマンドの作成
    subparsers = parser.add_subparsers(dest="command", help="サブコマンド")

    # entries サブコマンド
    parser_entries = subparsers.add_parser(
        "entries", help="エントリーポイント候補を表示"
    )
    parser_entries.add_argument(
        "--no-strict",
        action="store_false",
        dest="strict",
        help="緩和モード（デフォルトは厳密モード）",
    )
    parser_entries.add_argument(
        "--min-calls",
        type=int,
        default=1,
        help="エントリーポイントの最小呼び出し数 (デフォルト: 1)",
    )
    parser_entries.add_argument(
        "--tsv",
        action="store_true",
        help="TSV形式で出力",
    )

    # search サブコマンド
    parser_search = subparsers.add_parser("search", help="キーワードでメソッドを検索")
    parser_search.add_argument("keyword", help="検索キーワード")

    # forward サブコマンド
    parser_forward = subparsers.add_parser(
        "forward", help="指定メソッドからの呼び出しツリーを表示"
    )
    parser_forward.add_argument("method", help="起点メソッド")
    parser_forward.add_argument(
        "--depth", type=int, default=50, help="ツリーの最大深度 (デフォルト: 50)"
    )
    parser_forward.add_argument(
        "--show-class",
        action="store_true",
        dest="show_class",
        help="クラス情報を表示",
    )
    parser_forward.add_argument(
        "--show-sql", action="store_true", dest="show_sql", help="SQL情報を表示"
    )
    parser_forward.add_argument(
        "--no-follow-impl",
        action="store_false",
        dest="follow_impl",
        help="実装クラス候補を追跡しない",
    )
    parser_forward.add_argument(
        "--verbose",
        action="store_true",
        help="詳細表示（Javadocをタブ区切りで表示）",
    )
    parser_forward.add_argument(
        "--tab",
        action="store_true",
        help="ハードタブでインデントし、プレフィックス|-- を省略",
    )
    parser_forward.add_argument(
        "--short",
        action="store_true",
        help="クラス名からパッケージ名を省いて表示",
    )

    # reverse サブコマンド
    parser_reverse = subparsers.add_parser(
        "reverse", help="指定メソッドへの呼び出し元ツリーを表示"
    )
    parser_reverse.add_argument("method", help="対象メソッド")
    parser_reverse.add_argument(
        "--depth", type=int, default=50, help="ツリーの最大深度 (デフォルト: 50)"
    )
    parser_reverse.add_argument(
        "--show-class",
        action="store_true",
        dest="show_class",
        help="クラス情報を表示",
    )
    parser_reverse.add_argument(
        "--no-follow-override",
        action="store_false",
        dest="follow_override",
        help="オーバーライド元を追跡しない",
    )
    parser_reverse.add_argument(
        "--verbose",
        action="store_true",
        help="詳細表示（Javadocをタブ区切りで表示）",
    )
    parser_reverse.add_argument(
        "--tab",
        action="store_true",
        help="ハードタブでインデントし、プレフィックス|-- を省略",
    )
    parser_reverse.add_argument(
        "--short",
        action="store_true",
        help="クラス名からパッケージ名を省いて表示",
    )

    # export サブコマンド
    parser_export = subparsers.add_parser(
        "export", help="ツリーをファイルにエクスポート"
    )
    parser_export.add_argument("method", help="起点メソッド")
    parser_export.add_argument("output_file", help="出力ファイル名")
    parser_export.add_argument(
        "--format",
        choices=["text", "markdown", "html"],
        default="text",
        help="出力形式 (デフォルト: text)",
    )
    parser_export.add_argument(
        "--depth", type=int, default=50, help="ツリーの最大深度 (デフォルト: 50)"
    )
    parser_export.add_argument(
        "--no-follow-impl",
        action="store_false",
        dest="follow_impl",
        help="実装クラス候補を追跡しない",
    )

    # export-excel サブコマンド
    parser_export_excel = subparsers.add_parser(
        "export-excel", help="ツリーをExcelにエクスポート"
    )
    parser_export_excel.add_argument("output_file", help="出力Excelファイル名")
    parser_export_excel.add_argument(
        "--entry-points",
        help="エントリーポイントファイル（指定しない場合は厳密モードのエントリーポイントを使用）",
    )
    parser_export_excel.add_argument(
        "--depth", type=int, default=20, help="ツリーの最大深度 (デフォルト: 20)"
    )
    parser_export_excel.add_argument(
        "--no-follow-impl",
        action="store_false",
        dest="follow_impl",
        help="実装クラス候補を追跡しない",
    )
    parser_export_excel.add_argument(
        "--no-tree",
        action="store_false",
        dest="include_tree",
        help="L列以降の呼び出しツリーを出力しない",
    )
    parser_export_excel.add_argument(
        "--no-sql",
        action="store_false",
        dest="include_sql",
        help="SQL文を出力しない（動的列）",
    )

    # export-csv サブコマンド
    parser_export_csv = subparsers.add_parser(
        "export-csv", help="呼び出しメソッド一覧をCSVにエクスポート"
    )
    parser_export_csv.add_argument(
        "-o",
        "--output",
        dest="output_file",
        help="出力CSVファイル名（省略時は標準出力）",
    )
    parser_export_csv.add_argument(
        "--entry-points",
        help="エントリーポイントファイル（指定しない場合は厳密モードのエントリーポイントを使用）",
    )
    parser_export_csv.add_argument(
        "--depth", type=int, default=20, help="ツリーの最大深度 (デフォルト: 20)"
    )
    parser_export_csv.add_argument(
        "--no-follow-impl",
        action="store_false",
        dest="follow_impl",
        help="実装クラス候補を追跡しない",
    )

    # extract-sql サブコマンド
    parser_extract_sql = subparsers.add_parser(
        "extract-sql", help="SQL文を抽出してファイル出力"
    )
    parser_extract_sql.add_argument(
        "--output-dir",
        default="./found_sql",
        help="SQL出力先ディレクトリ (デフォルト: ./found_sql)",
    )

    # analyze-tables サブコマンド
    parser_analyze_tables = subparsers.add_parser(
        "analyze-tables", help="SQLファイルから使用テーブルを検出"
    )
    parser_analyze_tables.add_argument(
        "--sql-dir",
        default="./found_sql",
        help="SQLディレクトリ (デフォルト: ./found_sql)",
    )
    parser_analyze_tables.add_argument(
        "--table-list",
        default="./table_list.tsv",
        help="テーブルリストファイル (デフォルト: ./table_list.tsv)",
    )

    # class-tree サブコマンド
    parser_class_tree = subparsers.add_parser(
        "class-tree", help="クラス階層ツリーを表示"
    )
    parser_class_tree.add_argument(
        "--filter",
        help="フィルタリングパターン（パッケージ名やクラス名の一部）",
    )
    parser_class_tree.add_argument(
        "--root",
        dest="root_filter",
        help="ルートクラス指定（--filterのエイリアス）",
    )
    parser_class_tree.add_argument(
        "--verbose",
        action="store_true",
        help="詳細表示（Javadocやアノテーションを表示）",
    )

    # interface-impls サブコマンド
    parser_interface_impls = subparsers.add_parser(
        "interface-impls", help="インターフェース実装一覧を表示"
    )
    parser_interface_impls.add_argument(
        "--interface",
        dest="filter_str",
        help="特定のインターフェースでフィルタリング",
    )
    parser_interface_impls.add_argument(
        "--verbose",
        action="store_true",
        help="詳細表示（Javadocやアノテーションを表示）",
    )

    # 引数を解析
    args = parser.parse_args()

    # サブコマンドが指定されていない場合
    if not args.command:
        parser.print_help()
        sys.exit(1)

    # class-tree と interface-impls サブコマンドは CallTreeVisualizer を使わない
    if args.command == "class-tree":
        handle_class_tree(args)
        return
    elif args.command == "interface-impls":
        handle_interface_impls(args)
        return

    # Visualizerの初期化
    visualizer = CallTreeVisualizer(
        args.input_file, args.exclusion_file, args.output_tsv_encoding, args.debug_mode
    )

    # サブコマンドに応じた処理を実行
    if args.command == "entries":
        handle_entries(args, visualizer)
    elif args.command == "search":
        handle_search(args, visualizer)
    elif args.command == "forward":
        handle_forward(args, visualizer)
    elif args.command == "reverse":
        handle_reverse(args, visualizer)
    elif args.command == "export":
        handle_export(args, visualizer)
    elif args.command == "export-excel":
        handle_export_excel(args, visualizer)
    elif args.command == "export-csv":
        handle_export_csv(args, visualizer)
    elif args.command == "extract-sql":
        handle_extract_sql(args, visualizer)
    elif args.command == "analyze-tables":
        handle_analyze_tables(args, visualizer)


if __name__ == "__main__":
    main()
